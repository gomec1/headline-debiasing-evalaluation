"""
BiasScore Evaluation App meine Bachelorarbeit.

Die App vergleicht mehrere LLMs bei der Erkennung und Reformulierung von
Headlines. Es gibt vier Tabs: LLM-Konfiguration, Bewerter, Reformulierer
und Dashboard & Export. Die Prompts liegen als editierbare Textdateien im
Ordner prompts/. API-Keys werden getrennt in secrets.json gespeichert.
"""

# ============================================================
# ABSCHNITT 1: IMPORTE UND GLOBALE EINSTELLUNGEN
# ============================================================

import contextvars
import itertools
import json
import os
import re
import time
import traceback
from datetime import datetime
from hashlib import sha256
from pathlib import Path

import anthropic
import gradio as gr
import pandas as pd
from openai import APITimeoutError as OpenAITimeoutError
from openai import BadRequestError as OpenAIBadRequestError
from openai import OpenAI
from openai import RateLimitError as OpenAIRateLimitError
from sklearn.metrics import cohen_kappa_score, confusion_matrix, precision_recall_fscore_support

try:
    from google import genai
except ImportError:
    genai = None


BASE_DIR = Path(__file__).resolve().parent
CONFIG_PFAD = BASE_DIR / "config.json"
SECRETS_PFAD = BASE_DIR / "secrets.json"
PROMPT_ORDNER = BASE_DIR / "prompts"
EXPORT_ORDNER = BASE_DIR / "exports"
LOG_ORDNER = BASE_DIR / "logs"
CONFIG_EXAMPLE_PFAD = BASE_DIR / "config.example.json"

APP_VERSION = "2.0"
CSV_SEP = ";"
MAX_MINI = 50
RANDOM_STATE = 42

PROVIDER = ["OpenAI", "Anthropic", "Google (Gemini)", "Groq", "Together AI", "Ollama (lokal)"]
GEMINI_BASE_URL = "https://generativelanguage.googleapis.com/v1beta/openai/"
GROQ_BASE_URL = "https://api.groq.com/openai/v1"
TOGETHER_BASE_URL = "https://api.together.xyz/v1"
OLLAMA_BASE_URL = "http://localhost:11434/v1"
OLLAMA_RESPONSE_FORMAT_MODELS = set()
OPENAI_JSON_SCHEMA_PROVIDERS = {"OpenAI"}
OPENAI_JSON_OBJECT_PROVIDERS = {"OpenAI", "Google (Gemini)", "Groq", "Together AI"}
NATIVE_GEMINI_STRUCTURED_OUTPUT = True
GROQ_CACHING_SUPPORTED_MODELS = [
    "moonshotai/kimi-k2-instruct",
    "openai/gpt-oss-20b",
    "openai/gpt-oss-120b",
]

PROMPT_DATEIEN = {
    "linguistic_bias_bewerter": "bewerter_linguistic_bias.txt",
    "hyperpartisan_bewerter": "bewerter_hyperpartisan.txt",
    "linguistic_bias_reformulierer": "reformulierer_linguistic_bias.txt",
    "hyperpartisan_reformulierer": "reformulierer_hyperpartisan.txt",
}

ALTE_PROMPT_DATEIEN = {
    "bewerter_linguistic_bias.txt": "SYSTEM_PROMPT_BEWERTER_LINGUISTIC_BIAS.txt",
    "bewerter_hyperpartisan.txt": "SYSTEM_PROMPT_BEWERTER_HYPERPARTISAN.txt",
    "reformulierer_linguistic_bias.txt": "SYSTEM_PROMPT_REFORMULIERER_LINGUISTIC_BIAS.txt",
    "reformulierer_hyperpartisan.txt": "SYSTEM_PROMPT_REFORMULIERER_HYPERPARTISAN.txt",
}

# Globaler Session-Speicher. Modus A in Tab 3 liest daraus die Rohdaten aus Tab 2.
ERGEBNISSE = {
    "bewerter_2_1": None,
    "bewerter_2_2": None,
    "reformulierer_3_1": None,
    "reformulierer_3_2": None,
}

_PROMPT_CACHE = {}
_EMBED_MODELL = None
_CACHE_STATS = {}
_GROQ_CACHE_WARNED = set()
_AKTIVE_LOG_DATEI = contextvars.ContextVar("aktive_log_datei", default=None)
_AKTIVE_JSONL_LOG_DATEI = contextvars.ContextVar("aktive_jsonl_log_datei", default=None)
_AKTIVE_RUN_ID = contextvars.ContextVar("aktive_run_id", default=None)

JSON_PIPELINE_AUDIT_SCHEMA_VERSION = "1.0"
LOG_RAW_LLM_OUTPUTS = os.environ.get("LOG_RAW_LLM_OUTPUTS", "false").strip().lower() in {"1", "true", "yes", "ja"}
LOG_RAW_LLM_OUTPUT_LIMIT = int(os.environ.get("LOG_RAW_LLM_OUTPUT_LIMIT", "1000"))
SENSITIVE_LOG_KEYS = {
    "api_key", "apikey", "apiKey", "key", "token", "access_token", "refresh_token",
    "authorization", "bearer", "secret", "password", "api-key", "x-api-key",
}

BIAS_DIMENSIONS = {
    "linguistic_bias": ["framing", "intensifier", "verb", "labeling"],
    "hyperpartisan": ["emotional_tone", "one_sidedness", "conflict_framing", "identity_signaling"],
}

BIAS_EXPORT_FIELDS = {
    "linguistic_bias": [
        "category", "total_score", "framing", "framing_evidence", "intensifier", "intensifier_evidence",
        "verb", "verb_evidence", "labeling", "labeling_evidence", "reasoning",
    ],
    "hyperpartisan": [
        "category", "total_score", "binary_label", "emotional_tone", "emotional_tone_evidence",
        "one_sidedness", "one_sidedness_evidence", "conflict_framing", "conflict_framing_evidence",
        "identity_signaling", "identity_signaling_evidence", "reasoning",
    ],
}

BIAS_ERROR_FIELDS = ["status", "error_message", "provider", "model", "retry_count", "json_status", "json_warnings", "correction_applied", "raw_output_available"]

REFORMULIERER_FIELDS = [
    "neutralized_headline", "changed_terms", "meaning_preservation", "neutralization_summary", "changed_meaning_risk",
]


class JsonOutputError(ValueError):
    """Strukturierter Fehler fuer endgueltig ungueltige LLM-JSON-Ausgaben."""

    def __init__(self, message: str, *, raw_output: str = "", provider: str = "", model: str = "", retry_count: int = 0):
        super().__init__(message)
        self.raw_output = raw_output
        self.provider = provider
        self.model = model
        self.retry_count = retry_count


class AnalyseRunLogger:
    """Kleiner dateibasierter Logger fuer einen Analyse-Durchgang."""

    def __init__(self, path: Path, run_id: str | None = None, jsonl_path: Path | None = None):
        self.path = Path(path)
        self.run_id = run_id or self.path.stem
        self.jsonl_path = Path(jsonl_path) if jsonl_path else self.path.with_suffix(".jsonl")

    def write(self, message: str) -> None:
        timestamp_text = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        try:
            self.path.parent.mkdir(exist_ok=True)
            with open(self.path, "a", encoding="utf-8") as handle:
                for line in str(sanitize_for_logging(message)).splitlines() or [""]:
                    handle.write(f"[{timestamp_text}] {line}\n")
        except Exception as exc:
            print(f"[WARN] Konnte nicht in Analyse-Log schreiben ({self.path}): {sanitize_for_logging(str(exc))}")

    def write_json_event(self, event: dict) -> None:
        try:
            self.jsonl_path.parent.mkdir(exist_ok=True)
            event = sanitize_for_logging(event)
            with open(self.jsonl_path, "a", encoding="utf-8") as handle:
                handle.write(json.dumps(event, ensure_ascii=False, default=str) + "\n")
        except Exception as exc:
            print(f"[WARN] Konnte nicht in JSON-Pipeline-Log schreiben ({self.jsonl_path}): {sanitize_for_logging(str(exc))}")


def log_schreiben(message: str, path: Path | None = None, logger: AnalyseRunLogger | None = None) -> None:
    """Schreibt robust in eine explizite Logdatei oder als Fallback in die aktive ContextVar-Datei."""
    try:
        if logger is not None:
            logger.write(message)
            return
        ziel = path or _AKTIVE_LOG_DATEI.get()
        if ziel:
            AnalyseRunLogger(ziel).write(message)
    except Exception as exc:
        print(f"[WARN] Logging fehlgeschlagen: {sanitize_for_logging(str(exc))}")


def log_print(*args, logger: AnalyseRunLogger | None = None, log_path: Path | None = None, **kwargs) -> None:
    """Schreibt eine Logzeile ins Terminal und optional in die aktive Analyse-Logdatei."""
    safe_args = tuple(sanitize_for_logging(arg) for arg in args)
    print(*safe_args, **kwargs)
    sep = kwargs.get("sep", " ")
    end = kwargs.get("end", "\n")
    log_schreiben(sep.join(str(arg) for arg in safe_args) + end, path=log_path, logger=logger)


def _mask_secret_value(value) -> str:
    text = str(value)
    if len(text) <= 8:
        return "***REDACTED***"
    return f"{text[:4]}...{text[-4:]}***REDACTED***"


def _redact_secret_patterns(text: str) -> str:
    patterns = [
        r"(?i)(authorization\s*[:=]\s*bearer\s+)([^\s,;\"'}]+)",
        r"(?i)(bearer\s+)([A-Za-z0-9._\-]{16,})",
        r"sk-ant-api03-[A-Za-z0-9_\-]{20,}",
        r"sk-proj-[A-Za-z0-9_\-]{20,}",
        r"sk-[A-Za-z0-9_\-]{20,}",
        r"tgp_v1_[A-Za-z0-9_\-]{20,}",
        r"AIza[0-9A-Za-z_\-]{20,}",
    ]
    result = text
    for pattern in patterns:
        if pattern.startswith("(?i)(authorization") or pattern.startswith("(?i)(bearer"):
            result = re.sub(pattern, lambda match: match.group(1) + _mask_secret_value(match.group(2)), result)
        else:
            result = re.sub(pattern, lambda match: _mask_secret_value(match.group(0)), result)
    return result


def sanitize_for_logging(obj):
    """Redigiert Secrets rekursiv aus Strings, Dicts und Listen vor Logs/JSONL."""
    if isinstance(obj, dict):
        safe = {}
        for key, value in obj.items():
            key_text = str(key)
            if key_text.lower() in {s.lower() for s in SENSITIVE_LOG_KEYS} or any(token in key_text.lower() for token in ["api_key", "apikey", "token", "authorization", "secret", "password"]):
                safe[key] = _mask_secret_value(value)
            else:
                safe[key] = sanitize_for_logging(value)
        return safe
    if isinstance(obj, (list, tuple)):
        return [sanitize_for_logging(item) for item in obj]
    if isinstance(obj, set):
        return [sanitize_for_logging(item) for item in obj]
    if isinstance(obj, str):
        return _redact_secret_patterns(obj)
    if isinstance(obj, (int, float, bool)) or obj is None:
        return obj
    return _redact_secret_patterns(str(obj))


def _json_sicher(value):
    """Macht beliebige Werte JSONL-tauglich, ohne Pipeline-Logging abstuerzen zu lassen."""
    value = sanitize_for_logging(value)
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, dict):
        return {str(key): _json_sicher(val) for key, val in value.items()}
    if isinstance(value, (list, tuple, set)):
        return [_json_sicher(item) for item in value]
    try:
        json.dumps(value)
        return value
    except Exception:
        return sanitize_for_logging(str(value))


def prepare_raw_output_for_audit(value, field_name: str) -> dict:
    """Bereitet Raw-LLM-Output datensparsam und redigiert fuer Audit-Logs vor."""
    if value is None:
        return {f"{field_name}_available": False}
    original_text = value if isinstance(value, str) else json.dumps(value, ensure_ascii=False, default=str)
    text = sanitize_for_logging(original_text)
    result = {
        f"{field_name}_available": True,
        f"{field_name}_length_chars": len(original_text),
    }
    if LOG_RAW_LLM_OUTPUTS:
        result[f"{field_name}_truncated"] = False
        result[field_name] = text
    else:
        truncated = len(text) > LOG_RAW_LLM_OUTPUT_LIMIT
        result[f"{field_name}_truncated"] = truncated
        result[f"{field_name}_preview"] = text[:LOG_RAW_LLM_OUTPUT_LIMIT] if truncated else text
    return result


def _input_id_aus_text(text: str | None) -> str | None:
    if text is None:
        return None
    return sha256(str(text).encode("utf-8", errors="ignore")).hexdigest()[:16]


def record_json_pipeline_event(**event) -> None:
    """Schreibt ein strukturiertes JSON-Pipeline-Audit-Event in JSONL und Kurzform ins Run-Log."""
    try:
        raw_fields = {key: event.pop(key) for key in list(event.keys()) if key in {"raw_output_before_correction", "raw_output_first_attempt", "raw_output_last_attempt"}}
        event = {key: _json_sicher(value) for key, value in event.items() if value is not None}
        event.setdefault("timestamp", datetime.now().isoformat(timespec="seconds"))
        event.setdefault("run_id", _AKTIVE_RUN_ID.get())
        event.setdefault("audit_schema_version", JSON_PIPELINE_AUDIT_SCHEMA_VERSION)
        event["provider"] = event.get("provider") or "unknown"
        event["model"] = event.get("model") or "unknown"
        for raw_key, raw_value in raw_fields.items():
            event.update(prepare_raw_output_for_audit(raw_value, raw_key))
        if any(key.endswith("_available") and key.startswith("raw_output") for key in event):
            event["raw_output_available"] = any(value for key, value in event.items() if key.endswith("_available") and key.startswith("raw_output"))
            event["raw_output_truncated"] = any(value for key, value in event.items() if key.endswith("_truncated") and key.startswith("raw_output"))
        event = sanitize_for_logging(event)
        jsonl_path = _AKTIVE_JSONL_LOG_DATEI.get()
        if jsonl_path:
            AnalyseRunLogger(Path(jsonl_path).with_suffix(".log"), run_id=event.get("run_id"), jsonl_path=Path(jsonl_path)).write_json_event(event)
        if _AKTIVE_LOG_DATEI.get():
            stage = event.get("stage", "unknown_stage")
            status = event.get("status", "")
            correction = event.get("correction_type", "")
            model = event.get("model", "")
            log_print(f"[JSON_PIPELINE] stage={stage} status={status} correction={correction} model={model}")
    except Exception as exc:
        log_print(f"[WARN] JSON-Pipeline-Logging fehlgeschlagen: {sanitize_for_logging(str(exc))}")


def _json_context_correction(json_context: dict | None, correction_type: str) -> None:
    """Merkt Parse-/Repair-Korrekturen im laufenden JSON-Kontext fuer finale Metadaten."""
    if isinstance(json_context, dict):
        json_context["correction_applied"] = True
        corrections = json_context.setdefault("correction_types", [])
        if correction_type not in corrections:
            corrections.append(correction_type)


def analyse_log_starten(analyse_name: str) -> AnalyseRunLogger:
    """Legt eine neue Logdatei fuer einen Analyse-Durchgang an."""
    LOG_ORDNER.mkdir(exist_ok=True)
    safe_name = re.sub(r"[^a-zA-Z0-9_.-]+", "_", analyse_name).strip("_") or "analyse"
    run_id = f"{timestamp()}_{datetime.now().strftime('%S_%f')}_{safe_name}"
    path = LOG_ORDNER / f"{run_id}.log"
    jsonl_path = LOG_ORDNER / f"{run_id}_json_pipeline.jsonl"
    logger = AnalyseRunLogger(path, run_id=run_id, jsonl_path=jsonl_path)
    _AKTIVE_LOG_DATEI.set(path)
    _AKTIVE_JSONL_LOG_DATEI.set(jsonl_path)
    _AKTIVE_RUN_ID.set(run_id)
    log_print(f"[RUN] Start: {analyse_name}", logger=logger)
    log_print(f"[RUN] Logdatei: {path}", logger=logger)
    log_print(f"[RUN] JSON-Pipeline-Log: {jsonl_path}", logger=logger)
    return logger


def analyse_log_beenden(analyse_name: str, status: str, logger: AnalyseRunLogger) -> None:
    """Schliesst die aktive Analyse-Logdatei mit einem Status ab."""
    try:
        _AKTIVE_LOG_DATEI.set(logger.path)
        _AKTIVE_JSONL_LOG_DATEI.set(logger.jsonl_path)
        _AKTIVE_RUN_ID.set(logger.run_id)
        log_print(f"[RUN] Ende: {analyse_name} | Status: {status}", logger=logger)
        log_print(f"[RUN] Log gespeichert: {logger.path}", logger=logger)
        log_print(f"[RUN] JSON-Pipeline-Log gespeichert: {logger.jsonl_path}", logger=logger)
    except Exception as exc:
        safe_exc = sanitize_for_logging(str(exc))
        print(f"[WARN] Analyse-Logging konnte nicht sauber beendet werden ({logger.path}): {safe_exc}")
        log_schreiben(f"[WARN] Analyse-Logging konnte nicht sauber beendet werden: {safe_exc}", logger=logger)
    finally:
        _AKTIVE_LOG_DATEI.set(None)
        _AKTIVE_JSONL_LOG_DATEI.set(None)
        _AKTIVE_RUN_ID.set(None)


def _status_aus_update(output) -> str | None:
    """Extrahiert kompakt eine Statusmeldung aus einem Gradio-Generator-Update."""
    if isinstance(output, str):
        return output
    if isinstance(output, (tuple, list)) and output:
        for value in reversed(output):
            if isinstance(value, str) and value.strip():
                return value
    return None


def _update_mit_status(output, status_text: str):
    """Ersetzt die letzte String-Statuskomponente, ohne die Gradio-Output-Struktur zu aendern."""
    if isinstance(output, str):
        return status_text
    if isinstance(output, tuple):
        values = list(output)
        for index in range(len(values) - 1, -1, -1):
            if isinstance(values[index], str):
                values[index] = status_text
                return tuple(values)
        return output
    if isinstance(output, list):
        values = list(output)
        for index in range(len(values) - 1, -1, -1):
            if isinstance(values[index], str):
                values[index] = status_text
                return values
        return output
    return output


def analyse_mit_logging(analyse_name: str, generator_func, *args, **kwargs):
    """Fuehrt einen Analyse-Generator mit Datei-Logging und Abbrucherkennung aus."""
    logger = analyse_log_starten(analyse_name)
    path = logger.path
    final_status = "abgebrochen"
    first_status = True
    try:
        generator = iter(generator_func(*args, **kwargs))
        while True:
            _AKTIVE_LOG_DATEI.set(path)
            _AKTIVE_JSONL_LOG_DATEI.set(logger.jsonl_path)
            _AKTIVE_RUN_ID.set(logger.run_id)
            try:
                output = next(generator)
            except StopIteration:
                break
            status_text = _status_aus_update(output)
            if status_text:
                final_status = status_text
                ui_status = f"{status_text}\nLogdatei: {path}" if first_status else status_text
                log_print(f"[STATUS] {ui_status}", logger=logger)
                output = _update_mit_status(output, ui_status)
                first_status = False
            yield output
        if final_status == "abgebrochen":
            final_status = "fertig"
    except GeneratorExit:
        final_status = "manuell abgebrochen"
        log_print(f"[RUN] Abbruch angefordert: {analyse_name}", logger=logger)
        raise
    except Exception as exc:
        final_status = f"Fehler: {exc}"
        log_print(f"[ERROR] {exc}", logger=logger)
        log_schreiben(traceback.format_exc(), logger=logger)
        raise
    finally:
        analyse_log_beenden(analyse_name, final_status, logger)


# ============================================================
# ABSCHNITT 2: BASISDATEIEN, SECRETS UND KONFIGURATION
# ============================================================

def timestamp() -> str:
    """Erzeugt einen kurzen Zeitstempel fuer Exportdateien.

    Eingabe: keine.
    Ausgabe: Text wie 2026-05-17_1430.
    Wird von allen Exportfunktionen genutzt.
    """
    return datetime.now().strftime("%Y-%m-%d_%H%M")


def ensure_basic_files() -> None:
    """Legt technische Basisdateien an, falls sie fehlen.

    Eingabe: keine.
    Ausgabe: Ordner prompts/ und exports/ sowie config.example.json.
    Diese Funktion schreibt keine API-Keys.
    """
    PROMPT_ORDNER.mkdir(exist_ok=True)
    EXPORT_ORDNER.mkdir(exist_ok=True)
    if not CONFIG_EXAMPLE_PFAD.exists():
        write_json(CONFIG_EXAMPLE_PFAD, {
            "models": [{
                "name": "GPT-4o Mini",
                "provider": "OpenAI",
                "model_id": "gpt-4o-mini",
                "base_url": "",
            }]
        })


def read_json(path: Path, fallback):
    """Liest JSON fehlertolerant.

    Eingabe: Pfad und Ersatzwert.
    Ausgabe: Dateiinhalt oder Ersatzwert.
    So startet die App auch, wenn optionale Dateien fehlen.
    """
    if not path.exists():
        return fallback
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return fallback


def write_json(path: Path, data) -> None:
    """Schreibt JSON lesbar formatiert.

    Eingabe: Pfad und Daten.
    Ausgabe: keine.
    API-Keys werden nur ueber save_config() in secrets.json geschrieben.
    """
    path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")


def load_raw_config() -> list:
    """Liest config.json im alten oder neuen Format.

    Eingabe: keine.
    Ausgabe: Liste von Modell-Eintraegen.
    Altes Format war eine Liste, neues Format ist {"models": [...]}.
    """
    raw = read_json(CONFIG_PFAD, {"models": []})
    if isinstance(raw, list):
        return raw
    if isinstance(raw, dict):
        return raw.get("models", []) or []
    return []


def normalize_model_config(entry: dict, secrets: dict | None = None) -> dict:
    """Vereinheitlicht alte und neue Modell-Konfigurationen.

    Eingabe: Eintrag aus config.json und optional secrets.json.
    Ausgabe: internes Dict mit model_id, modell_id und api_key.
    Dadurch bleibt alter Code-Stil kompatibel.
    """
    secrets = secrets or {}
    name = str(entry.get("name", "")).strip()
    model_id = str(entry.get("model_id") or entry.get("modell_id") or "").strip()
    return {
        "name": name,
        "provider": str(entry.get("provider") or "OpenAI").strip(),
        "model_id": model_id,
        "modell_id": model_id,
        "api_key": str(entry.get("api_key") or secrets.get(name, "")).strip(),
        "base_url": str(entry.get("base_url") or "").strip(),
    }


def migrate_config_if_needed() -> None:
    """Verschiebt alte Klartext-Keys automatisch nach secrets.json.

    Eingabe: keine.
    Ausgabe: bereinigte config.json und secrets.json, falls Migration noetig ist.
    Im Terminal wird nur die Anzahl Keys ausgegeben, nie der Key selbst.
    """
    raw_json = read_json(CONFIG_PFAD, {"models": []})
    raw_models = load_raw_config()
    if not raw_models:
        return
    has_old_keys = any(isinstance(m, dict) and "api_key" in m for m in raw_models)
    has_old_field = any(isinstance(m, dict) and "modell_id" in m for m in raw_models)
    if not has_old_keys and not has_old_field and not isinstance(raw_json, list):
        return

    secrets = read_json(SECRETS_PFAD, {})
    cleaned = []
    moved = 0
    for entry in raw_models:
        if not isinstance(entry, dict):
            continue
        cfg = normalize_model_config(entry)
        if cfg["name"] and cfg["api_key"]:
            secrets[cfg["name"]] = cfg["api_key"]
            moved += 1
        cleaned.append({
            "name": cfg["name"],
            "provider": cfg["provider"],
            "model_id": cfg["model_id"],
            "base_url": cfg["base_url"],
        })
    write_json(CONFIG_PFAD, {"models": cleaned})
    write_json(SECRETS_PFAD, secrets)
    print(f"Migration: {moved} Keys nach secrets.json verschoben, config.json bereinigt.")


def load_config() -> list[dict]:
    """Laedt LLM-Konfigurationen plus Secrets.

    Eingabe: keine.
    Ausgabe: kombinierte interne Modellliste.
    API-Keys bleiben intern nutzbar, werden aber nicht in Tabellen angezeigt.
    """
    migrate_config_if_needed()
    secrets = read_json(SECRETS_PFAD, {})
    return [normalize_model_config(m, secrets) for m in load_raw_config() if isinstance(m, dict)]


def save_config(models: list[dict]) -> None:
    """Speichert config.json und secrets.json getrennt.

    Eingabe: interne Modellliste.
    Ausgabe: config.json ohne Keys und secrets.json mit Keys.
    Das UI bleibt gleich, aber Git bekommt keine Secrets.
    """
    cleaned = []
    secrets = read_json(SECRETS_PFAD, {})
    for entry in models:
        cfg = normalize_model_config(entry)
        if not cfg["name"]:
            continue
        cleaned.append({
            "name": cfg["name"],
            "provider": cfg["provider"],
            "model_id": cfg["model_id"],
            "base_url": cfg["base_url"],
        })
        if cfg["api_key"]:
            secrets[cfg["name"]] = cfg["api_key"]
        else:
            secrets.pop(cfg["name"], None)
    write_json(CONFIG_PFAD, {"models": cleaned})
    write_json(SECRETS_PFAD, secrets)


def default_base_url(provider: str, base_url: str) -> str:
    """Setzt Standard-Basis-URLs, wenn das Feld leer ist.

    Eingabe: Provider und UI-Wert.
    Ausgabe: passende Basis-URL oder leer.
    Groq und Gemini brauchen dadurch keine manuelle URL.
    """
    if (base_url or "").strip():
        return base_url.strip()
    provider_key = str(provider or "").strip().lower()
    if provider == "Google (Gemini)":
        return GEMINI_BASE_URL
    if provider == "Groq":
        return GROQ_BASE_URL
    if provider == "Together AI" or provider_key in {"together", "together.ai", "together-ai"}:
        return TOGETHER_BASE_URL
    if provider == "Ollama (lokal)":
        return OLLAMA_BASE_URL
    return ""


def llm_namen() -> list[str]:
    """Gibt alle gespeicherten LLM-Namen zurueck.

    Eingabe: keine.
    Ausgabe: Liste der Anzeigenamen.
    Diese Liste fuellt alle LLM-Auswahlfelder.
    """
    return [c["name"] for c in load_config() if c.get("name")]


def config_finden(name: str) -> dict | None:
    """Sucht eine LLM-Konfiguration nach Anzeigename.

    Eingabe: Name aus der UI.
    Ausgabe: Konfiguration oder None.
    Alle Analysen holen so ihre Modell-Daten.
    """
    for cfg in load_config():
        if cfg.get("name") == name:
            return cfg
    return None


def tabelle_aus_config() -> pd.DataFrame:
    """Erzeugt eine sichere Konfigurations-Tabelle.

    Eingabe: keine.
    Ausgabe: DataFrame ohne API-Key-Werte.
    Fehlende Keys werden sichtbar markiert.
    """
    rows = []
    for cfg in load_config():
        key_status = "lokal/optional" if cfg["provider"] == "Ollama (lokal)" else ("vorhanden" if cfg.get("api_key") else "FEHLT")
        rows.append({
            "Name": cfg["name"],
            "Provider": cfg["provider"],
            "Modell-ID": cfg["model_id"],
            "Basis-URL": cfg.get("base_url", ""),
            "API-Key": key_status,
        })
    return pd.DataFrame(rows, columns=["Name", "Provider", "Modell-ID", "Basis-URL", "API-Key"])


def dropdown_updates():
    """Aktualisiert alle LLM-Auswahlfelder.

    Eingabe: keine.
    Ausgabe: gr.update-Objekte fuer vier Multiselects und Entfernen-Dropdown.
    Nach Speichern/Loeschen ist die UI wieder synchron.
    """
    names = llm_namen()
    return (
        gr.update(choices=names, value=names),
        gr.update(choices=names, value=names),
        gr.update(choices=names, value=names),
        gr.update(choices=names, value=names),
        gr.update(choices=names, value=names[0] if names else None),
    )


def llm_hinzufuegen(name, provider, model_id, api_key, base_url):
    """Speichert oder ersetzt ein LLM aus Tab 1.

    Eingabe: UI-Felder.
    Ausgabe: Tabelle, Status und aktualisierte Dropdowns.
    Der API-Key wird nach secrets.json geschrieben.
    """
    name = (name or "").strip()
    model_id = (model_id or "").strip()
    if not name or not model_id:
        return (tabelle_aus_config(), "Fehler: Name und Modell-ID muessen ausgefuellt sein.", *dropdown_updates())
    if provider == "Groq" and model_id not in GROQ_CACHING_SUPPORTED_MODELS:
        print(
            f"[INFO] Prompt Caching fuer Groq-Modell '{model_id}' nicht verfuegbar. "
            f"Caching wird unterstuetzt von: {', '.join(GROQ_CACHING_SUPPORTED_MODELS)}"
        )
    models = [m for m in load_config() if m.get("name") != name]
    models.append({
        "name": name,
        "provider": provider,
        "model_id": model_id,
        "api_key": (api_key or "").strip(),
        "base_url": default_base_url(provider, base_url or ""),
    })
    save_config(models)
    return (tabelle_aus_config(), f"LLM '{name}' gespeichert.", *dropdown_updates())


def llm_entfernen(name):
    """Entfernt ein LLM und seinen Secret-Eintrag.

    Eingabe: Modellname.
    Ausgabe: Tabelle, Status und aktualisierte Dropdowns.
    Dadurch bleiben keine verwaisten Keys liegen.
    """
    models = [m for m in load_config() if m.get("name") != name]
    secrets = read_json(SECRETS_PFAD, {})
    secrets.pop(name, None)
    write_json(SECRETS_PFAD, secrets)
    save_config(models)
    return (tabelle_aus_config(), f"LLM '{name}' entfernt.", *dropdown_updates())


# ============================================================
# ABSCHNITT 3: PROMPTS LADEN UND BEREINIGEN
# ============================================================

def extract_prompt_text(text: str) -> str:
    """Entfernt Python-Wrapper aus alten Prompt-Dateien.

    Eingabe: Dateiinhalt, eventuell VARIABLE = triple quotes.
    Ausgabe: reiner Prompt-Text.
    Danach kann der Prompt direkt editiert werden.
    """
    match = re.search(r'=\s*[rRuU]?("""|\'\'\')(.*)\1\s*$', text, re.DOTALL)
    if match:
        return match.group(2).strip() + "\n"
    return text.strip() + "\n"


def migrate_prompt_files() -> None:
    """Erstellt neue reine Prompt-Dateien aus vorhandenen alten Dateien.

    Eingabe: keine.
    Ausgabe: vier neue Textdateien, falls sie fehlen.
    Existierende neue Dateien werden nicht ueberschrieben.
    """
    PROMPT_ORDNER.mkdir(exist_ok=True)
    for new_name, old_name in ALTE_PROMPT_DATEIEN.items():
        new_path = PROMPT_ORDNER / new_name
        old_path = PROMPT_ORDNER / old_name
        if new_path.exists() or not old_path.exists():
            continue
        new_path.write_text(extract_prompt_text(old_path.read_text(encoding="utf-8")), encoding="utf-8")


def load_prompt(filename: str) -> str:
    """Laedt einen Prompt aus prompts/.

    Eingabe: Dateiname.
    Ausgabe: Prompt-Text.
    Bei fehlender Datei entsteht eine klare UI-Fehlermeldung.
    """
    if filename in _PROMPT_CACHE:
        return _PROMPT_CACHE[filename]
    path = PROMPT_ORDNER / filename
    if not path.exists():
        raise FileNotFoundError(f"Datei `prompts/{filename}` nicht gefunden. Bitte erstellen.")
    text = path.read_text(encoding="utf-8")
    _PROMPT_CACHE[filename] = text
    return text


def reload_prompts():
    """Leert den Prompt-Cache.

    Eingabe: keine.
    Ausgabe: Statusmeldung.
    Der naechste LLM-Call liest die Prompt-Dateien frisch ein.
    """
    _PROMPT_CACHE.clear()
    return "Prompts wurden neu geladen. Der naechste LLM-Call nutzt die aktuellen Dateien."


def prompt_for(prompt_type: str, role: str) -> str:
    """Waehlt den passenden Prompt.

    Eingabe: Bias-Art und Rolle bewerter/reformulierer.
    Ausgabe: Prompt-Text.
    So bleibt die Prompt-Auswahl zentral und nachvollziehbar.
    """
    return load_prompt(PROMPT_DATEIEN[f"{prompt_type}_{role}"])


# ============================================================
# ABSCHNITT 4: LLM-AUFRUF UND JSON-NORMALISIERUNG
# ============================================================

def _status_melden(status_callback, message: str, label: str = "RATE LIMIT") -> None:
    """Meldet Wartezeiten an die UI, falls ein Callback uebergeben wurde.

    Eingabe: optionale Callback-Funktion und Text.
    Ausgabe: keine.
    Ohne Callback bleibt die Funktion rueckwaertskompatibel.
    """
    message = sanitize_for_logging(message)
    log_print(f"[{label}] {message}")
    # Gradio zeigt gr.Info als kleine Meldung im Browser. Der optionale
    # Callback bleibt trotzdem da, damit Generator-Funktionen die Meldung
    # zusaetzlich in ihr Status-Markdown uebernehmen koennen.
    try:
        gr.Info(message)
    except Exception:
        pass
    if status_callback:
        status_callback(message)


def _header_wert(headers, name: str):
    """Liest einen Header fehlertolerant aus.

    Eingabe: Header-Objekt und Headername.
    Ausgabe: Headerwert oder None.
    Nicht jeder Provider sendet dieselben Rate-Limit-Header.
    """
    if not headers:
        return None
    try:
        return headers.get(name)
    except Exception:
        return None


def _sekunden_aus_header(value, fallback: float) -> float:
    """Wandelt Header-Zeitangaben in Sekunden um.

    Eingabe: Werte wie "7.66s", "30" oder None.
    Ausgabe: Sekunden als float.
    Falls der Header fehlt oder unlesbar ist, wird der Fallback genutzt.
    """
    if value is None:
        return fallback
    text_value = str(value).strip().lower()
    if text_value.endswith("s"):
        text_value = text_value[:-1]
    try:
        return max(float(text_value), 0.0)
    except ValueError:
        return fallback


def _rate_limit_sleep_nach_erfolg(headers, model_name: str, status_callback=None) -> None:
    """Wartet proaktiv, wenn laut Response kaum Requests uebrig sind.

    Eingabe: Response-Header, Modellname und optionaler Status-Callback.
    Ausgabe: keine.
    Das verhindert oft den naechsten HTTP-429-Fehler.
    """
    remaining = _header_wert(headers, "x-ratelimit-remaining-requests")
    if remaining is not None:
        try:
            remaining_number = float(str(remaining).strip())
        except ValueError:
            remaining_number = None
        if remaining_number is not None and remaining_number <= 1:
            wait_time = _sekunden_aus_header(_header_wert(headers, "x-ratelimit-reset-tokens"), 10.0)
            _status_melden(status_callback, f"Rate Limit bei {model_name}: nur noch {int(remaining_number)} Request(s) uebrig - warte {wait_time:.1f}s vor dem naechsten Call.")
            time.sleep(wait_time)
    # Kleiner Standard-Puffer nach jedem erfolgreichen Call gegen Burst-Probleme.
    time.sleep(0.5)


def _retry_after_aus_exception(exc, fallback: float) -> float:
    """Ermittelt die Wartezeit nach einem HTTP-429-Fehler.

    Eingabe: SDK-Exception und Fallback-Wartezeit.
    Ausgabe: Sekunden.
    Groq setzt retry-after bei 429; andere Provider oft nicht.
    """
    response = getattr(exc, "response", None)
    headers = getattr(response, "headers", None)
    return _sekunden_aus_header(_header_wert(headers, "retry-after"), fallback)


def reset_cache_stats() -> None:
    """Setzt die Cache-Zaehler fuer eine neue Analyse zurueck.

    Eingabe: keine.
    Ausgabe: keine.
    So zeigt die Zusammenfassung nur die aktuelle Analyse-Session.
    """
    _CACHE_STATS.clear()


def _cache_stats_erhoehen(provider: str, cache_read: int = 0, cache_write: int = 0) -> None:
    """Zaehlt gelesene und geschriebene Cache-Tokens.

    Eingabe: Providername und Tokenzahlen.
    Ausgabe: keine.
    Die Werte werden am Ende einer Analyse im Terminal ausgegeben.
    """
    stats = _CACHE_STATS.setdefault(provider, {"cache_read": 0, "cache_write": 0})
    stats["cache_read"] += int(cache_read or 0)
    stats["cache_write"] += int(cache_write or 0)


def print_cache_summary(cache_stats: dict | None = None) -> None:
    """Zeigt die Cache-Statistik am Ende einer Analyse.

    Eingabe: optionales Statistik-Dict.
    Ausgabe: Terminal-Ausgabe.
    Die Zahlen helfen zu pruefen, ob Prompt Caching wirklich greift.
    """
    cache_stats = cache_stats if cache_stats is not None else _CACHE_STATS
    log_print("\n=== Cache-Statistik ===")
    any_value = False
    for provider, stats in cache_stats.items():
        read = int(stats.get("cache_read", 0) or 0)
        write = int(stats.get("cache_write", 0) or 0)
        if read > 0 or write > 0:
            any_value = True
            log_print(f"{provider}: {read:,} Tokens aus Cache | {write:,} Tokens gecacht")
    if not any_value:
        log_print("Keine Cache-Hits oder Cache-Writes gemeldet.")
    log_print("======================\n")


def _prompt_cache_stats_loggen(provider: str, model_id: str, response) -> None:
    """Liest Cache-Informationen aus Provider-Responses.

    Eingabe: Provider, Modell-ID und SDK-Response.
    Ausgabe: Debug-Ausgabe und aktualisierte Statistik.
    OpenAI/Gemini/Groq cachen automatisch; Anthropic liefert Read/Write-Werte.
    """
    # PROMPT CACHING:
    # Der System-Prompt ist bei vielen Calls gleich. Einige Anbieter koennen
    # diesen Prompt wiederverwenden, statt ihn jedes Mal neu zu verarbeiten.
    # Das spart Kosten. Hier lesen wir nur die gemeldeten Cache-Zahlen aus.
    cache_read = 0
    cache_write = 0

    usage = getattr(response, "usage", None)
    if provider == "Anthropic":
        cache_read = int(getattr(usage, "cache_read_input_tokens", 0) or 0)
        cache_write = int(getattr(usage, "cache_creation_input_tokens", 0) or 0)
        if cache_read > 0:
            log_print(f"[DEBUG] Anthropic Cache-HIT: {cache_read} Tokens aus Cache gelesen (gespart!)")
        elif cache_write > 0:
            log_print(f"[DEBUG] Anthropic Cache geschrieben: {cache_write} Tokens gecacht fuer naechste Calls")
    elif provider == "Google (Gemini)":
        usage_metadata = getattr(response, "usage_metadata", None)
        cache_read = int(getattr(usage_metadata, "cached_content_token_count", 0) or 0)
        if cache_read == 0 and usage is not None:
            details = getattr(usage, "prompt_tokens_details", None)
            cache_read = int(getattr(details, "cached_tokens", 0) or 0)
        if cache_read > 0:
            log_print(f"[DEBUG] Gemini Cache-HIT: {cache_read} Tokens aus Cache gelesen (gespart!)")
    elif provider == "Groq":
        if model_id not in GROQ_CACHING_SUPPORTED_MODELS and model_id not in _GROQ_CACHE_WARNED:
            _GROQ_CACHE_WARNED.add(model_id)
            log_print(
                f"[INFO] Prompt Caching fuer Groq-Modell '{model_id}' nicht verfuegbar. "
                f"Caching wird unterstuetzt von: {', '.join(GROQ_CACHING_SUPPORTED_MODELS)}"
            )
        details = getattr(usage, "prompt_tokens_details", None)
        cache_read = int(getattr(details, "cached_tokens", 0) or 0)
        if cache_read > 0:
            log_print(f"[DEBUG] Groq Cache-HIT: {cache_read} Tokens aus Cache gelesen (gespart!)")
    elif provider == "OpenAI":
        details = getattr(usage, "prompt_tokens_details", None)
        cache_read = int(getattr(details, "cached_tokens", 0) or 0)
        if cache_read > 0:
            log_print(f"[DEBUG] OpenAI Cache-HIT: {cache_read} Tokens aus Cache gelesen (gespart!)")

    if cache_read > 0 or cache_write > 0:
        _cache_stats_erhoehen(provider, cache_read=cache_read, cache_write=cache_write)


def llm_bewerten(text: str, modell_config: dict, system_prompt: str, status_callback=None, output_schema: dict | None = None, force_json: bool = False, prefer_json_schema: bool = False) -> str:
    """Fuehrt genau einen LLM-Call aus.

    Eingabe: User-Text, Modell-Konfiguration, System-Prompt und optionaler Status-Callback.
    Ausgabe: Rohantwort als Text.
    Diese Funktion ist die einzige Stelle fuer externe LLM-Aufrufe.
    """
    if modell_config is None:
        raise ValueError("LLM-Konfiguration nicht gefunden.")
    model_name = modell_config.get("name") or modell_config.get("model_id") or "Unbekanntes Modell"
    start = time.monotonic()
    log_print(f"[DEBUG] Starte LLM-Call: {model_name} | Eingabe: {text[:60]}")

    # ============================================================
    # RATE-LIMIT-HANDLING
    # ============================================================
    # Manche API-Anbieter erlauben nur eine bestimmte Anzahl Anfragen
    # pro Minute (RPM = Requests Per Minute). Groq zum Beispiel erlaubt
    # auf dem Free Plan nur 30 Anfragen/Minute. Weil Groq extrem schnell
    # antwortet, koennen diese 30 Anfragen in wenigen Sekunden verbraucht
    # sein. Dann kommt HTTP 429 ("Too Many Requests").
    #
    # Unsere Loesung:
    # 1. Nach jedem erfolgreichen Call pruefen wir Header wie
    #    x-ratelimit-remaining-requests. Wenn fast nichts uebrig ist,
    #    warten wir proaktiv.
    # 2. Wenn ein 429 kommt, lesen wir retry-after. Dieser Header sagt,
    #    wie viele Sekunden wir warten sollen.
    # 3. Wenn retry-after fehlt, warten wir 30/60/120 Sekunden
    #    (Exponential Backoff).
    # 4. Nach maximal 3 Versuchen geben wir fuer diese Headline auf.
    # ============================================================

    backoff_seconds = [30.0, 60.0, 120.0]

    # NUR FUER TESTS:
    # Um Rate Limits ohne viele echte Requests zu pruefen, kann man hier
    # temporaer eine simulierte 429-Situation einbauen, z. B. wenn
    # os.environ.get("DEBUG_RATE_LIMIT") == "1". Der Block bleibt bewusst
    # auskommentiert, damit normale Analysen nie kuenstlich warten.
    #
    # if os.environ.get("DEBUG_RATE_LIMIT") == "1":
    #     _status_melden(status_callback, f"Rate Limit bei {model_name} - warte 1.0s (Debug-Test).")
    #     time.sleep(1.0)

    if modell_config["provider"] == "Anthropic":
        client = anthropic.Anthropic(api_key=modell_config.get("api_key", ""), timeout=90.0)
        for attempt in range(3):
            try:
                raw_resp = client.messages.with_raw_response.create(
                    model=modell_config["model_id"],
                    max_tokens=1000,
                    system=system_prompt,
                    messages=[{"role": "user", "content": text}],
                )
                resp = raw_resp.parse()
                _rate_limit_sleep_nach_erfolg(raw_resp.headers, model_name, status_callback)
                log_print(f"[DEBUG] Antwort erhalten: {model_name} | Zeit: {time.monotonic() - start:.1f}s")
                return resp.content[0].text
            except anthropic.RateLimitError as exc:
                if attempt == 2:
                    raise RuntimeError("Rate Limit erschöpft - Headline übersprungen") from exc
                wait_time = _retry_after_aus_exception(exc, backoff_seconds[attempt])
                _status_melden(status_callback, f"Rate Limit bei {model_name} - warte {wait_time:.1f}s (Versuch {attempt + 1}/3).")
                time.sleep(wait_time)
            except (anthropic.APITimeoutError, TimeoutError) as exc:
                raise TimeoutError(f"Timeout: {model_name} hat nach 90s nicht geantwortet.") from exc

    base_url = modell_config.get("base_url") or None
    provider_key = str(modell_config["provider"] or "").strip().lower()
    if modell_config["provider"] == "Google (Gemini)" and not base_url:
        base_url = GEMINI_BASE_URL
    if modell_config["provider"] == "Groq" and not base_url:
        base_url = GROQ_BASE_URL
    if (modell_config["provider"] == "Together AI" or provider_key in {"together", "together.ai", "together-ai"}) and not base_url:
        # Together AI nutzt die OpenAI-kompatible Chat-API, aber mit eigener Basis-URL.
        # Ohne diese URL wuerde der tgp_v1-Key faelschlich an OpenAI geschickt.
        base_url = TOGETHER_BASE_URL
    if modell_config["provider"] == "Ollama (lokal)" and not base_url:
        base_url = OLLAMA_BASE_URL
    client = OpenAI(api_key=modell_config.get("api_key") or "ollama", base_url=base_url, timeout=90.0)
    request_kwargs = {}
    is_json_prompt = re.search(r"\bjson\b", system_prompt or "", flags=re.IGNORECASE)
    is_ollama = modell_config["provider"] == "Ollama (lokal)"
    ollama_json_enabled = modell_config["model_id"] in OLLAMA_RESPONSE_FORMAT_MODELS
    supports_json_schema = bool(modell_config.get("supports_json_schema")) or (modell_config["provider"] in OPENAI_JSON_SCHEMA_PROVIDERS and not base_url)
    supports_json_object = bool(modell_config.get("supports_json_object")) or (modell_config["provider"] in OPENAI_JSON_OBJECT_PROVIDERS and not is_ollama) or (is_ollama and ollama_json_enabled)
    if output_schema and prefer_json_schema and supports_json_schema:
        request_kwargs["response_format"] = {
            "type": "json_schema",
            "json_schema": {
                "name": "bias_evaluation_result",
                "strict": True,
                "schema": output_schema,
            },
        }
    elif (force_json or is_json_prompt) and supports_json_object:
        request_kwargs["response_format"] = {"type": "json_object"}
    for attempt in range(3):
        try:
            raw_resp = client.chat.completions.with_raw_response.create(
                model=modell_config["model_id"],
                temperature=0,
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": text},
                ],
                **request_kwargs,
            )
            resp = raw_resp.parse()
            _rate_limit_sleep_nach_erfolg(raw_resp.headers, model_name, status_callback)
            log_print(f"[DEBUG] Antwort erhalten: {model_name} | Zeit: {time.monotonic() - start:.1f}s")
            return resp.choices[0].message.content or ""
        except OpenAIRateLimitError as exc:
            if attempt == 2:
                raise RuntimeError("Rate Limit erschöpft - Headline übersprungen") from exc
            wait_time = _retry_after_aus_exception(exc, backoff_seconds[attempt])
            _status_melden(status_callback, f"Rate Limit bei {model_name} - warte {wait_time:.1f}s (Versuch {attempt + 1}/3).")
            time.sleep(wait_time)
        except OpenAIBadRequestError as exc:
            response_format = request_kwargs.get("response_format") or {}
            if response_format.get("type") == "json_schema":
                log_print(f"[WARN] JSON-Schema-Mode von {model_name} abgelehnt, wechsle auf JSON-Object-Mode: {exc}")
                request_kwargs["response_format"] = {"type": "json_object"}
                continue
            raise
        except (OpenAITimeoutError, TimeoutError) as exc:
            raise TimeoutError(f"Timeout: {model_name} hat nach 90s nicht geantwortet.") from exc

    raise RuntimeError("Rate Limit erschöpft - Headline übersprungen")


def bewerter_model_output_schema(prompt_type: str) -> dict:
    """Erzeugt das JSON-Schema fuer die primaere Modellantwort ohne abgeleitete Felder."""
    dimensions = BIAS_DIMENSIONS[prompt_type]
    properties = {dimension: {"type": "integer", "minimum": 0, "maximum": 3} for dimension in dimensions}
    properties.update({
        "dimension_evidence": {
            "type": "object",
            "additionalProperties": False,
            "properties": {dimension: {"type": "string"} for dimension in dimensions},
            "required": dimensions,
        },
        "reasoning": {"type": "string"},
    })
    return {"type": "object", "additionalProperties": False, "properties": properties, "required": [*dimensions, "dimension_evidence", "reasoning"]}


def normalized_result_schema(prompt_type: str) -> dict:
    """Erzeugt das interne Schema fuer gespeicherte normalisierte Bewerter-Ergebnisse."""
    schema = bewerter_model_output_schema(prompt_type)
    properties = dict(schema["properties"])
    properties.update({
        "total_score": {"type": "integer", "minimum": 0, "maximum": 12},
        "category": {"type": "string", "enum": ["Low", "Medium", "High"]},
    })
    if prompt_type == "hyperpartisan":
        properties["binary_label"] = {"type": "string", "enum": ["non-hyperpartisan", "hyperpartisan"]}
    required = [*BIAS_DIMENSIONS[prompt_type], "dimension_evidence", "reasoning", "total_score", "category"]
    if prompt_type == "hyperpartisan":
        required.append("binary_label")
    return {"type": "object", "additionalProperties": False, "properties": properties, "required": required}


def rewriter_input_schema(prompt_type: str) -> dict:
    """Schema fuer den Reformulierer-Input: Original plus normalisierte Bewerteranalyse."""
    return {
        "type": "object",
        "additionalProperties": False,
        "properties": {
            "original_headline": {"type": "string"},
            "bias_analysis": normalized_result_schema(prompt_type),
        },
        "required": ["original_headline", "bias_analysis"],
    }


def model_output_schema_bewerter(prompt_type: str) -> dict:
    return bewerter_model_output_schema(prompt_type)


def normalized_result_schema_bewerter(prompt_type: str) -> dict:
    return normalized_result_schema(prompt_type)


def rewriter_model_output_schema() -> dict:
    """Schema fuer Reformulierer-Antworten ohne Bewerter- oder Score-Felder."""
    return {
        "type": "object",
        "additionalProperties": False,
        "properties": {
            "neutralized_headline": {"type": "string"},
            "changed_terms": {
                "type": "array",
                "items": {
                    "type": "object",
                    "additionalProperties": False,
                    "properties": {
                        "dimension": {"type": "string"},
                        "original": {"type": "string"},
                        "replacement": {"type": "string"},
                        "reason": {"type": "string"},
                    },
                    "required": ["dimension", "original", "replacement", "reason"],
                },
            },
            "meaning_preservation": {"type": "string"},
            "neutralization_summary": {"type": "string"},
            "changed_meaning_risk": {"type": "string", "enum": ["low", "medium", "high"]},
        },
        "required": ["neutralized_headline", "changed_terms", "meaning_preservation", "neutralization_summary", "changed_meaning_risk"],
    }


def model_output_schema_rewriter() -> dict:
    return rewriter_model_output_schema()


def bewerter_json_schema(prompt_type: str) -> dict:
    """Kompatibilitaets-Alias: Schema fuer die Modellantwort."""
    return bewerter_model_output_schema(prompt_type)


def _score_zu_category_en(score: int) -> str:
    if score <= 3:
        return "Low"
    if score <= 7:
        return "Medium"
    return "High"


def _schema_als_text(schema: dict) -> str:
    return json.dumps(schema, ensure_ascii=False, indent=2)


def _schema_fuer_gemini(schema: dict) -> dict:
    """Wandelt das lokale JSON-Schema in Geminis response_schema-Dialekt um."""
    type_map = {"object": "OBJECT", "integer": "INTEGER", "string": "STRING", "array": "ARRAY", "boolean": "BOOLEAN", "number": "NUMBER"}
    converted = {}
    schema_type = schema.get("type")
    if schema_type:
        converted["type"] = type_map.get(schema_type, schema_type)
    if "enum" in schema:
        converted["enum"] = schema["enum"]
    if "properties" in schema:
        converted["properties"] = {key: _schema_fuer_gemini(value) for key, value in schema["properties"].items()}
        converted["propertyOrdering"] = list(schema["properties"].keys())
    if "required" in schema:
        converted["required"] = schema["required"]
    if "items" in schema:
        converted["items"] = _schema_fuer_gemini(schema["items"])
    return converted


def _score_feld_validieren(data: dict, field: str, warnings: list[str]) -> int:
    if field not in data:
        raise ValueError(f"Pflichtfeld '{field}' fehlt im JSON")
    value = data[field]
    if isinstance(value, bool):
        raise ValueError(f"Score-Feld '{field}' muss ein Integer 0-3 sein, erhalten: {value!r}")
    if isinstance(value, int):
        score = value
    elif isinstance(value, str) and re.fullmatch(r"[0-3]", value.strip()):
        score = int(value.strip())
        warnings.append(f"Score-Feld '{field}' war String und wurde in Integer umgewandelt.")
    else:
        raise ValueError(f"Score-Feld '{field}' muss ein Integer 0-3 sein, erhalten: {value!r}")
    if not 0 <= score <= 3:
        raise ValueError(f"Score-Feld '{field}' liegt ausserhalb 0-3: {score!r}")
    return score


def _total_score_lesen(value, warnings: list[str]) -> int | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value if 0 <= value <= 12 else None
    if isinstance(value, str) and re.fullmatch(r"\d{1,2}", value.strip()):
        warnings.append("total_score war String und wurde nur fuer den Vergleich in Integer umgewandelt.")
        parsed = int(value.strip())
        return parsed if 0 <= parsed <= 12 else None
    return None


def _binary_label_normalisieren(value) -> str:
    s = str(value or "").strip().lower()
    if s in {"1", "true", "yes", "hyperpartisan", "biased"}:
        return "hyperpartisan"
    if s in {"0", "false", "no", "non-hyperpartisan", "nonhyperpartisan", "neutral", "unbiased"}:
        return "non-hyperpartisan"
    raise ValueError(f"Unbekanntes binary_label: {value!r}")


def compute_derived_fields(data: dict, prompt_type: str, warnings: list[str] | None = None) -> dict:
    """Berechnet total_score, category und ggf. binary_label deterministisch lokal."""
    warnings = warnings if warnings is not None else []
    result = {key: value for key, value in data.items() if key not in {"total_score", "gesamt", "category", "kategorie", "binary_label"}}
    dimensions = BIAS_DIMENSIONS[prompt_type]
    total_score = sum(int(result[dimension]) for dimension in dimensions)
    local_category = _score_zu_category_en(total_score)

    if "total_score" in data:
        model_total = _total_score_lesen(data.get("total_score"), warnings)
        if model_total != total_score:
            warnings.append(f"total_score vom Modell ({data.get('total_score')!r}) wurde ignoriert; lokale Summe ist {total_score}.")
        else:
            warnings.append("total_score vom Modell wurde ignoriert und lokal neu berechnet.")
    if "category" in data:
        model_category = str(data.get("category") or "").strip()
        if model_category.lower() != local_category.lower():
            warnings.append(f"category vom Modell ({data.get('category')!r}) wurde ignoriert; lokale Kategorie ist {local_category!r}.")
        else:
            warnings.append("category vom Modell wurde ignoriert und lokal neu berechnet.")

    result["total_score"] = total_score
    result["gesamt"] = total_score
    result["category"] = local_category
    result["kategorie"] = local_category

    if prompt_type == "hyperpartisan":
        local_binary = "non-hyperpartisan" if local_category == "Low" else "hyperpartisan"
        if "binary_label" in data:
            try:
                model_binary = _binary_label_normalisieren(data.get("binary_label"))
                if model_binary != local_binary:
                    warnings.append(f"binary_label vom Modell ({data.get('binary_label')!r}) wurde ignoriert; lokale Ableitung ist {local_binary!r}.")
                else:
                    warnings.append("binary_label vom Modell wurde ignoriert und lokal neu berechnet.")
            except ValueError:
                warnings.append(f"binary_label vom Modell ({data.get('binary_label')!r}) war ungueltig und wurde ignoriert; lokale Ableitung ist {local_binary!r}.")
        result["binary_label"] = local_binary
    return result


def validate_and_normalize_bewerter_json(data, prompt_type: str, modell=None, headline=None, json_context: dict | None = None) -> dict:
    """Validiert Bewerter-JSON zentral und gibt nur saubere, normalisierte Daten zurueck."""
    context = json_context or {}
    if not isinstance(data, dict):
        record_json_pipeline_event(stage="schema_validation_failed", status="failed", validation_errors=["Bewerter-Antwort muss ein JSON-Objekt sein"], parsed_output_before_normalization=data, **context)
        raise ValueError("Bewerter-Antwort muss ein JSON-Objekt sein")
    dimensions = BIAS_DIMENSIONS[prompt_type]
    derived_fields = {"total_score", "category", "binary_label"}
    allowed = {*dimensions, "dimension_evidence", "reasoning", *derived_fields}
    warnings = []
    unexpected = sorted(set(data) - allowed)
    if unexpected:
        warnings.append("Unerwartete Felder entfernt: " + ", ".join(unexpected))
    present_derived = sorted(set(data) & derived_fields)
    if present_derived:
        warnings.append("Abgeleitete Modellfelder werden ignoriert und lokal berechnet: " + ", ".join(present_derived))

    result = {}
    for dimension in dimensions:
        result[dimension] = _score_feld_validieren(data, dimension, warnings)

    evidence = data.get("dimension_evidence")
    if not isinstance(evidence, dict):
        raise ValueError("Pflichtfeld 'dimension_evidence' muss ein Objekt sein")
    missing_evidence = [dimension for dimension in dimensions if dimension not in evidence]
    if missing_evidence:
        raise ValueError("dimension_evidence fehlt fuer: " + ", ".join(missing_evidence))
    extra_evidence = sorted(set(evidence) - set(dimensions))
    if extra_evidence:
        warnings.append("Unerwartete dimension_evidence-Felder entfernt: " + ", ".join(extra_evidence))
    result["dimension_evidence"] = {}
    for dimension in dimensions:
        value = evidence[dimension]
        if value is None:
            value = ""
        if not isinstance(value, str):
            warnings.append(f"dimension_evidence.{dimension} wurde in String umgewandelt.")
            value = str(value)
        result["dimension_evidence"][dimension] = value
        result[f"{dimension}_evidence"] = value

    reasoning = data.get("reasoning")
    if not isinstance(reasoning, str):
        raise ValueError("Pflichtfeld 'reasoning' muss ein String sein")
    result["reasoning"] = reasoning.strip()

    derived_source = {field: data[field] for field in derived_fields if field in data}
    result = compute_derived_fields({**result, **derived_source}, prompt_type, warnings)

    if warnings:
        result["validation_warnings"] = warnings
        for warning in warnings:
            modell_text = modell or "unbekanntes Modell"
            headline_text = headline or "unbekannte Headline"
            log_print(f"[WARN] JSON-Validation fuer {modell_text} / {headline_text}: {warning}")
    correction_details = []
    if unexpected:
        correction_details.append({"correction_type": "unexpected_fields_removed", "removed_fields": unexpected})
        record_json_pipeline_event(stage="unexpected_fields_removed", status="corrected", correction_type="unexpected_fields_removed", correction_details=correction_details[-1:], parsed_output_before_normalization=data, normalized_output_after_correction=result, **context)
    if present_derived:
        model_values = {field: data.get(field) for field in present_derived}
        final_values = {field: result.get(field) for field in present_derived if field in result}
        correction_details.append({"correction_type": "unexpected_model_derived_fields", "model_values": model_values, "computed_values": final_values})
        record_json_pipeline_event(stage="model_fields_overwritten", status="corrected", correction_type="unexpected_model_derived_fields", correction_details=correction_details[-1:], parsed_output_before_normalization=data, normalized_output_after_correction=result, **context)
    for dimension in dimensions:
        if isinstance(data.get(dimension), str) and data.get(dimension).strip() != str(result.get(dimension)):
            pass
        if isinstance(data.get(dimension), str) and data.get(dimension).strip() == str(result.get(dimension)):
            correction_details.append({"correction_type": "score_type_normalized", "field": dimension, "before": data.get(dimension), "after": result.get(dimension), "reason": "score string converted to integer"})
    if "total_score" in data:
        correction_details.append({"correction_type": "derived_field_overwritten", "field": "total_score", "before": data.get("total_score"), "after": result.get("total_score"), "reason": "total_score computed locally from dimensions"})
    if "category" in data:
        correction_details.append({"correction_type": "derived_field_overwritten", "field": "category", "before": data.get("category"), "after": result.get("category"), "reason": "category computed locally from total_score"})
    if "binary_label" in data and prompt_type == "hyperpartisan":
        correction_details.append({"correction_type": "derived_field_overwritten", "field": "binary_label", "before": data.get("binary_label"), "after": result.get("binary_label"), "reason": "binary_label computed locally from category"})
    if extra_evidence:
        correction_details.append({"correction_type": "unexpected_fields_removed", "field": "dimension_evidence", "removed_fields": extra_evidence})
    record_json_pipeline_event(stage="derived_fields_computed", status="corrected" if correction_details else "valid", correction_type="derived_fields_computed", parsed_output_before_normalization=data, normalized_output_after_correction={"total_score": result.get("total_score"), "category": result.get("category"), "binary_label": result.get("binary_label")}, correction_details=[detail for detail in correction_details if detail.get("correction_type") == "derived_field_overwritten"], **context)
    if correction_details:
        record_json_pipeline_event(stage="normalization_applied", status="corrected", correction_type="normalization_applied", parsed_output_before_normalization=data, normalized_output_after_correction=result, correction_details=correction_details, validation_errors=warnings, **context)
    record_json_pipeline_event(stage="schema_validation_succeeded", status="corrected" if correction_details else "valid", parsed_output_before_normalization=data, normalized_output_after_correction=result, validation_errors=[], **context)
    result["json_status"] = "corrected" if correction_details else "valid"
    result["json_warnings"] = warnings
    result["correction_applied"] = bool(correction_details)
    result["raw_output_available"] = bool(context.get("raw_output_available", True))
    return result


def _bewerter_model_view(data: dict, prompt_type: str) -> dict:
    """Extrahiert aus gespeicherten Bewerterdaten nur die fachlichen Felder fuer Re-Normalisierung."""
    data = data or {}
    dimensions = BIAS_DIMENSIONS[prompt_type]
    view = {dimension: data.get(dimension) for dimension in dimensions if dimension in data}
    evidence = data.get("dimension_evidence") if isinstance(data.get("dimension_evidence"), dict) else {}
    view["dimension_evidence"] = {dimension: evidence.get(dimension, data.get(f"{dimension}_evidence", "")) for dimension in dimensions}
    view["reasoning"] = data.get("reasoning", "")
    return view


def ensure_normalized_bias_analysis(data: dict, prompt_type: str, modell=None, headline=None, json_context: dict | None = None) -> dict:
    """Backfillt alte Bewerter-Ergebnisse vor der Uebergabe an den Reformulierer."""
    return validate_and_normalize_bewerter_json(_bewerter_model_view(data or {}, prompt_type), prompt_type, modell=modell, headline=headline, json_context=json_context)


def validate_and_normalize_rewriter_json(data, modell=None, headline=None, json_context: dict | None = None) -> dict:
    """Validiert Reformulierer-JSON strikt getrennt vom Bewerter-Schema."""
    context = json_context or {}
    if not isinstance(data, dict):
        record_json_pipeline_event(stage="schema_validation_failed", status="failed", validation_errors=["Reformulierer-Antwort muss ein JSON-Objekt sein"], parsed_output_before_normalization=data, **context)
        raise ValueError("Reformulierer-Antwort muss ein JSON-Objekt sein")
    forbidden = set(BIAS_DIMENSIONS["linguistic_bias"] + BIAS_DIMENSIONS["hyperpartisan"] + [
        "total_score", "gesamt", "category", "kategorie", "binary_label", "dimension_evidence", "reasoning",
    ])
    forbidden_present = sorted(set(data) & forbidden)
    if forbidden_present:
        record_json_pipeline_event(stage="schema_validation_failed", status="failed", correction_type="rewriter_score_fields_detected", correction_details=[{"invalid_fields": forbidden_present, "action": "rejected"}], validation_errors=["Reformulierer-Output enthaelt Bewerter-/Score-Felder"], parsed_output_before_normalization=data, **context)
        raise ValueError("Reformulierer-Output enthaelt Bewerter-/Score-Felder: " + ", ".join(forbidden_present))

    required = ["neutralized_headline", "changed_terms", "meaning_preservation", "neutralization_summary", "changed_meaning_risk"]
    missing = [field for field in required if field not in data]
    if missing:
        record_json_pipeline_event(stage="schema_validation_failed", status="failed", validation_errors=["Reformulierer-Pflichtfelder fehlen: " + ", ".join(missing)], parsed_output_before_normalization=data, **context)
        raise ValueError("Reformulierer-Pflichtfelder fehlen: " + ", ".join(missing))
    warnings = []
    unexpected = sorted(set(data) - set(required))
    if unexpected:
        warnings.append("Unerwartete Reformulierer-Felder entfernt: " + ", ".join(unexpected))

    result = {}
    for field in ["neutralized_headline", "meaning_preservation", "neutralization_summary"]:
        value = data.get(field)
        if not isinstance(value, str):
            raise ValueError(f"Reformulierer-Feld '{field}' muss ein String sein")
        result[field] = value.strip()

    changed_terms = data.get("changed_terms")
    if not isinstance(changed_terms, list):
        raise ValueError("Reformulierer-Feld 'changed_terms' muss eine Liste sein")
    normalized_terms = []
    term_required = ["dimension", "original", "replacement", "reason"]
    for index, term in enumerate(changed_terms):
        if not isinstance(term, dict):
            raise ValueError(f"changed_terms[{index}] muss ein Objekt sein")
        missing_term = [field for field in term_required if field not in term]
        if missing_term:
            raise ValueError(f"changed_terms[{index}] fehlt: " + ", ".join(missing_term))
        extra_term = sorted(set(term) - set(term_required))
        if extra_term:
            warnings.append(f"Unerwartete Felder in changed_terms[{index}] entfernt: " + ", ".join(extra_term))
        normalized_term = {}
        for field in term_required:
            value = term.get(field)
            if not isinstance(value, str):
                raise ValueError(f"changed_terms[{index}].{field} muss ein String sein")
            normalized_term[field] = value.strip()
        normalized_terms.append(normalized_term)
    result["changed_terms"] = normalized_terms

    risk = str(data.get("changed_meaning_risk") or "").strip().lower()
    if risk not in {"low", "medium", "high"}:
        record_json_pipeline_event(stage="schema_validation_failed", status="failed", validation_errors=[f"changed_meaning_risk muss low, medium oder high sein: {data.get('changed_meaning_risk')!r}"], parsed_output_before_normalization=data, **context)
        raise ValueError(f"changed_meaning_risk muss low, medium oder high sein: {data.get('changed_meaning_risk')!r}")
    result["changed_meaning_risk"] = risk

    if warnings:
        result["validation_warnings"] = warnings
        for warning in warnings:
            modell_text = modell or "unbekanntes Modell"
            headline_text = headline or "unbekannte Headline"
            log_print(f"[WARN] Reformulierer-Validation fuer {modell_text} / {headline_text}: {warning}")
    correction_details = []
    if unexpected:
        correction_details.append({"correction_type": "unexpected_fields_removed", "removed_fields": unexpected})
    if str(data.get("changed_meaning_risk") or "").strip() != risk:
        correction_details.append({"correction_type": "enum_value_normalized", "field": "changed_meaning_risk", "before": data.get("changed_meaning_risk"), "after": risk})
    if correction_details:
        record_json_pipeline_event(stage="normalization_applied", status="corrected", correction_type="rewriter_normalization_applied", parsed_output_before_normalization=data, normalized_output_after_correction=result, correction_details=correction_details, validation_errors=warnings, **context)
    record_json_pipeline_event(stage="schema_validation_succeeded", status="corrected" if correction_details else "valid", parsed_output_before_normalization=data, normalized_output_after_correction=result, validation_errors=[], **context)
    result["json_status"] = "corrected" if correction_details else "valid"
    result["json_warnings"] = warnings
    result["correction_applied"] = bool(correction_details)
    result["raw_output_available"] = bool(context.get("raw_output_available", True))
    return result


def _build_rewriter_repair_prompt(schema: dict, raw_output: str, error_message: str) -> str:
    return f"""Your previous response was not valid or schema-compliant JSON.

Return exactly one valid JSON object matching the required rewriter output schema.
Do not use markdown.
Do not wrap the JSON in code fences.
Do not add text before or after the JSON.
Do not add fields.
Do not omit fields.
All string values must be valid JSON strings.
Escape quotation marks inside string values.
The output must not include score fields.
The output must not include total_score.
The output must not include category.
The output must not include binary_label.
The output must not include dimension_evidence.
The output must not include reasoning.

Required rewriter output schema:
{_schema_als_text(schema)}

Invalid previous response:
{raw_output}

Validation error:
{error_message}

Return only the corrected JSON object."""


def _build_json_repair_prompt(original_input: str, schema: dict, raw_output: str, error_message: str) -> str:
    return f"""Your previous response was not valid or schema-compliant JSON.

Return exactly one valid JSON object matching the required model output schema.
Do not use markdown.
Do not wrap the JSON in code fences.
Do not add text before or after the JSON.
Do not add fields.
Do not omit fields.
All score values must be integers from 0 to 3.
All string values must be valid JSON strings.
Escape quotation marks inside string values.
If evidence contains quotation marks from the headline, either escape them correctly or omit the quotation marks.
Do not include total_score in your response.
Do not include category in your response.
Do not include binary_label in your response.
The final explanation must be short and must not include chain-of-thought.

Original input:
{original_input}

Required model output schema:
{_schema_als_text(schema)}

Invalid previous response:
{raw_output}

Validation error:
{error_message}

Return only the corrected JSON object."""


def _gemini_native_json_call(text: str, modell_config: dict, system_prompt: str, output_schema: dict, status_callback=None) -> dict | str:
    if genai is None:
        raise RuntimeError("google-genai ist nicht installiert")
    model_name = modell_config.get("name") or modell_config.get("model_id") or "Gemini"
    start = time.monotonic()
    log_print(f"[DEBUG] Starte nativen Gemini-JSON-Call: {model_name} | Eingabe: {text[:60]}")
    client = genai.Client(api_key=modell_config.get("api_key", ""))
    response = client.models.generate_content(
        model=modell_config["model_id"],
        contents=text,
        config={
            "system_instruction": system_prompt,
            "temperature": 0,
            "response_mime_type": "application/json",
            "response_schema": _schema_fuer_gemini(output_schema),
        },
    )
    log_print(f"[DEBUG] Native Gemini-Antwort erhalten: {model_name} | Zeit: {time.monotonic() - start:.1f}s")
    parsed = getattr(response, "parsed", None)
    if parsed is not None:
        return parsed
    return getattr(response, "text", "") or ""


def call_llm_for_json(modell_config: dict, system_prompt: str, user_input: str, prompt_type: str, output_schema: dict, max_json_retries: int = 2, status_callback=None) -> dict:
    """Fuehrt einen LLM-Call fuer schema-konformes Bewerter-JSON mit Retry/Repair aus."""
    model_name = (modell_config or {}).get("name") or (modell_config or {}).get("model_id") or "Unbekanntes Modell"
    provider = (modell_config or {}).get("provider") or "unknown"
    raw_output = ""
    first_raw_output = ""
    last_error = ""
    use_native_gemini = provider == "Google (Gemini)" and NATIVE_GEMINI_STRUCTURED_OUTPUT and genai is not None and not modell_config.get("disable_native_structured_output")

    for retry_count in range(max_json_retries + 1):
        repair_mode = retry_count > 0
        prompt = system_prompt
        text = user_input
        if repair_mode:
            record_json_pipeline_event(stage="retry_started", status="retrying", correction_type="retry_triggered", provider=provider, model=model_name, pipeline_type="evaluator", prompt_type=prompt_type, input_id=_input_id_aus_text(user_input), original_headline=user_input, raw_output_before_correction=raw_output, validation_errors=[last_error], retry_count=retry_count)
            prompt = system_prompt + "\n\n" + _build_json_repair_prompt(user_input, output_schema, raw_output, last_error)
            text = "Return only the corrected JSON object."
            _status_melden(status_callback, f"JSON-Reparatur fuer {provider}/{model_name} (Versuch {retry_count}/{max_json_retries}) ...", label="JSON-REPAIR")
        try:
            json_context = {"provider": provider, "model": model_name, "pipeline_type": "evaluator", "prompt_type": prompt_type, "input_id": _input_id_aus_text(user_input), "original_headline": user_input, "retry_count": retry_count}
            if use_native_gemini:
                try:
                    response_payload = _gemini_native_json_call(text, modell_config, prompt, output_schema, status_callback=status_callback)
                    raw_output = json.dumps(response_payload, ensure_ascii=False) if isinstance(response_payload, dict) else str(response_payload)
                    if not first_raw_output:
                        first_raw_output = raw_output
                    record_json_pipeline_event(stage="raw_model_output_received", status="received", raw_output_before_correction=raw_output, **json_context)
                    if isinstance(response_payload, dict):
                        parsed = response_payload
                        record_json_pipeline_event(stage="json_parse_succeeded", status="valid", parsed_output_before_normalization=parsed, **json_context)
                    else:
                        parsed = json_aus_text(raw_output, json_context=json_context)
                except Exception as exc:
                    log_print(f"[WARN] Nativer Gemini-Structured-Output fehlgeschlagen, Fallback auf OpenAI-kompatiblen JSON-Mode: {exc}")
                    use_native_gemini = False
                    raw_output = llm_bewerten(text, modell_config, prompt, status_callback=status_callback, output_schema=output_schema, force_json=True, prefer_json_schema=False)
                    if not first_raw_output:
                        first_raw_output = raw_output
                    record_json_pipeline_event(stage="raw_model_output_received", status="received", raw_output_before_correction=raw_output, **json_context)
                    parsed = json_aus_text(raw_output, json_context=json_context)
            else:
                prefer_schema = provider == "OpenAI" or bool(modell_config.get("supports_json_schema"))
                raw_output = llm_bewerten(text, modell_config, prompt, status_callback=status_callback, output_schema=output_schema, force_json=True, prefer_json_schema=prefer_schema)
                if not first_raw_output:
                    first_raw_output = raw_output
                record_json_pipeline_event(stage="raw_model_output_received", status="received", raw_output_before_correction=raw_output, **json_context)
                parsed = json_aus_text(raw_output, json_context=json_context)
            normalized = validate_and_normalize_bewerter_json(parsed, prompt_type, modell=model_name, headline=user_input, json_context=json_context)
            normalized["json_retry_count"] = retry_count
            normalized["retry_count"] = retry_count
            normalized["json_output_mode"] = "native_gemini_schema" if provider == "Google (Gemini)" and use_native_gemini else ("json_schema_or_json_object" if provider == "OpenAI" else "json_object_or_prompt")
            if json_context.get("correction_applied"):
                normalized["correction_applied"] = True
                if normalized.get("json_status") == "valid":
                    normalized["json_status"] = "corrected"
            if retry_count > 0:
                normalized["json_status"] = "retry_corrected" if normalized.get("json_status") in {"valid", "corrected"} else normalized.get("json_status")
                normalized["correction_applied"] = True
                record_json_pipeline_event(stage="retry_succeeded", status="corrected", correction_type="retry_succeeded", raw_output_first_attempt=first_raw_output, raw_output_last_attempt=raw_output, normalized_output_after_correction=normalized, retry_count=retry_count, **{key: value for key, value in json_context.items() if key != "retry_count"})
            record_json_pipeline_event(stage="final_output_valid", status=normalized.get("json_status", "valid"), raw_output_before_correction=raw_output, parsed_output_before_normalization=parsed, normalized_output_after_correction=normalized, retry_count=retry_count, **{key: value for key, value in json_context.items() if key != "retry_count"})
            return normalized
        except Exception as exc:
            last_error = str(exc)
            json_context = {"provider": provider, "model": model_name, "pipeline_type": "evaluator", "prompt_type": prompt_type, "input_id": _input_id_aus_text(user_input), "original_headline": user_input, "retry_count": retry_count}
            record_json_pipeline_event(stage="schema_validation_failed", status="retryable" if retry_count < max_json_retries else "failed", validation_errors=[last_error], raw_output_before_correction=raw_output, **json_context)
            log_print(f"[WARN] JSON-Output ungueltig fuer {provider}/{model_name} (Retry {retry_count}/{max_json_retries}): {last_error}")
            if retry_count >= max_json_retries:
                record_json_pipeline_event(stage="retry_failed", status="failed", correction_type="retry_exhausted", raw_output_first_attempt=first_raw_output, raw_output_last_attempt=raw_output, validation_errors=[last_error], retry_count=retry_count, **{key: value for key, value in json_context.items() if key != "retry_count"})
                record_json_pipeline_event(stage="final_output_invalid", status="failed", final_status="failed", error_type="retry_exhausted", raw_output_first_attempt=first_raw_output, raw_output_last_attempt=raw_output, validation_errors=[last_error], retry_count=retry_count, **{key: value for key, value in json_context.items() if key != "retry_count"})
                raise JsonOutputError(
                    f"JSON-Validierung fehlgeschlagen nach {retry_count} Retry(s): {last_error}",
                    raw_output=raw_output,
                    provider=provider,
                    model=model_name,
                    retry_count=retry_count,
                ) from exc

    raise JsonOutputError("JSON-Validierung fehlgeschlagen", raw_output=raw_output, provider=provider, model=model_name, retry_count=max_json_retries)


def call_llm_for_rewriter_json(modell_config: dict, system_prompt: str, user_input: str, output_schema: dict, max_json_retries: int = 2, status_callback=None, headline=None, prompt_type: str = "") -> dict:
    """Fuehrt einen LLM-Call fuer schema-konformes Reformulierer-JSON mit eigenem Retry/Repair aus."""
    model_name = (modell_config or {}).get("name") or (modell_config or {}).get("model_id") or "Unbekanntes Modell"
    provider = (modell_config or {}).get("provider") or "unknown"
    raw_output = ""
    first_raw_output = ""
    last_error = ""
    use_native_gemini = provider == "Google (Gemini)" and NATIVE_GEMINI_STRUCTURED_OUTPUT and genai is not None and not modell_config.get("disable_native_structured_output")

    for retry_count in range(max_json_retries + 1):
        prompt = system_prompt
        text = user_input
        if retry_count > 0:
            record_json_pipeline_event(stage="retry_started", status="retrying", correction_type="retry_triggered", provider=provider, model=model_name, pipeline_type="rewriter", prompt_type=prompt_type, input_id=_input_id_aus_text(headline or user_input), original_headline=headline, raw_output_before_correction=raw_output, validation_errors=[last_error], retry_count=retry_count)
            prompt = system_prompt + "\n\n" + _build_rewriter_repair_prompt(output_schema, raw_output, last_error)
            text = "Return only the corrected JSON object."
            _status_melden(status_callback, f"JSON-Reparatur fuer Reformulierer {provider}/{model_name} (Versuch {retry_count}/{max_json_retries}) ...", label="JSON-REPAIR")
        try:
            json_context = {"provider": provider, "model": model_name, "pipeline_type": "rewriter", "prompt_type": prompt_type, "input_id": _input_id_aus_text(headline or user_input), "original_headline": headline, "retry_count": retry_count}
            if use_native_gemini:
                try:
                    response_payload = _gemini_native_json_call(text, modell_config, prompt, output_schema, status_callback=status_callback)
                    raw_output = json.dumps(response_payload, ensure_ascii=False) if isinstance(response_payload, dict) else str(response_payload)
                    if not first_raw_output:
                        first_raw_output = raw_output
                    record_json_pipeline_event(stage="raw_model_output_received", status="received", raw_output_before_correction=raw_output, **json_context)
                    if isinstance(response_payload, dict):
                        parsed = response_payload
                        record_json_pipeline_event(stage="json_parse_succeeded", status="valid", parsed_output_before_normalization=parsed, **json_context)
                    else:
                        parsed = json_aus_text(raw_output, json_context=json_context)
                except Exception as exc:
                    log_print(f"[WARN] Nativer Gemini-Structured-Output fuer Reformulierer fehlgeschlagen, Fallback auf OpenAI-kompatiblen JSON-Mode: {exc}")
                    use_native_gemini = False
                    raw_output = llm_bewerten(text, modell_config, prompt, status_callback=status_callback, output_schema=output_schema, force_json=True, prefer_json_schema=False)
                    if not first_raw_output:
                        first_raw_output = raw_output
                    record_json_pipeline_event(stage="raw_model_output_received", status="received", raw_output_before_correction=raw_output, **json_context)
                    parsed = json_aus_text(raw_output, json_context=json_context)
            else:
                prefer_schema = provider == "OpenAI" or bool(modell_config.get("supports_json_schema"))
                raw_output = llm_bewerten(text, modell_config, prompt, status_callback=status_callback, output_schema=output_schema, force_json=True, prefer_json_schema=prefer_schema)
                if not first_raw_output:
                    first_raw_output = raw_output
                record_json_pipeline_event(stage="raw_model_output_received", status="received", raw_output_before_correction=raw_output, **json_context)
                parsed = json_aus_text(raw_output, json_context=json_context)
            result = validate_and_normalize_rewriter_json(parsed, modell=model_name, headline=headline, json_context=json_context)
            result["json_retry_count"] = retry_count
            result["retry_count"] = retry_count
            result["json_output_mode"] = "native_gemini_schema" if provider == "Google (Gemini)" and use_native_gemini else ("json_schema_or_json_object" if provider == "OpenAI" else "json_object_or_prompt")
            if json_context.get("correction_applied"):
                result["correction_applied"] = True
                if result.get("json_status") == "valid":
                    result["json_status"] = "corrected"
            if retry_count > 0:
                result["json_status"] = "retry_corrected" if result.get("json_status") in {"valid", "corrected"} else result.get("json_status")
                result["correction_applied"] = True
                record_json_pipeline_event(stage="retry_succeeded", status="corrected", correction_type="retry_succeeded", raw_output_first_attempt=first_raw_output, raw_output_last_attempt=raw_output, normalized_output_after_correction=result, retry_count=retry_count, **{key: value for key, value in json_context.items() if key != "retry_count"})
            record_json_pipeline_event(stage="final_output_valid", status=result.get("json_status", "valid"), raw_output_before_correction=raw_output, parsed_output_before_normalization=parsed, normalized_output_after_correction=result, retry_count=retry_count, **{key: value for key, value in json_context.items() if key != "retry_count"})
            return result
        except Exception as exc:
            last_error = str(exc)
            json_context = {"provider": provider, "model": model_name, "pipeline_type": "rewriter", "prompt_type": prompt_type, "input_id": _input_id_aus_text(headline or user_input), "original_headline": headline, "retry_count": retry_count}
            record_json_pipeline_event(stage="schema_validation_failed", status="retryable" if retry_count < max_json_retries else "failed", validation_errors=[last_error], raw_output_before_correction=raw_output, **json_context)
            log_print(f"[WARN] Reformulierer-JSON ungueltig fuer {provider}/{model_name} (Retry {retry_count}/{max_json_retries}): {last_error}")
            if retry_count >= max_json_retries:
                record_json_pipeline_event(stage="retry_failed", status="failed", correction_type="retry_exhausted", raw_output_first_attempt=first_raw_output, raw_output_last_attempt=raw_output, validation_errors=[last_error], retry_count=retry_count, **{key: value for key, value in json_context.items() if key != "retry_count"})
                record_json_pipeline_event(stage="final_output_invalid", status="failed", final_status="failed", error_type="retry_exhausted", raw_output_first_attempt=first_raw_output, raw_output_last_attempt=raw_output, validation_errors=[last_error], retry_count=retry_count, **{key: value for key, value in json_context.items() if key != "retry_count"})
                raise JsonOutputError(
                    f"Reformulierer-JSON-Validierung fehlgeschlagen nach {retry_count} Retry(s): {last_error}",
                    raw_output=raw_output,
                    provider=provider,
                    model=model_name,
                    retry_count=retry_count,
                ) from exc

    raise JsonOutputError("Reformulierer-JSON-Validierung fehlgeschlagen", raw_output=raw_output, provider=provider, model=model_name, retry_count=max_json_retries)


def json_aus_text(text: str, json_context: dict | None = None) -> dict:
    """Extrahiert JSON aus einer LLM-Antwort.

    Eingabe: Rohtext vom Modell.
    Ausgabe: Python-Dict.
    <think>-Bloecke werden entfernt, weil manche Modelle Denktext mitschicken.
    """
    context = json_context or {}
    raw_text = text or ""
    record_json_pipeline_event(stage="json_extraction_attempted", status="started", raw_output_before_correction=raw_text, **context)
    cleaned = raw_text
    without_think = re.sub(r"<think>.*?</think>", "", cleaned, flags=re.DOTALL).strip()
    if without_think != cleaned.strip():
        _json_context_correction(context, "think_block_removed")
        record_json_pipeline_event(stage="think_block_removed", status="corrected", correction_type="think_block_removed", raw_output_before_correction=cleaned, normalized_output_after_correction=without_think, **context)
    cleaned = without_think
    without_fence = re.sub(r"^```(?:json)?\s*", "", cleaned, flags=re.IGNORECASE).strip()
    without_fence = re.sub(r"\s*```$", "", without_fence).strip()
    if without_fence != cleaned.strip():
        _json_context_correction(context, "markdown_fence_removed")
        record_json_pipeline_event(stage="markdown_fence_removed", status="corrected", correction_type="markdown_fence_removed", raw_output_before_correction=cleaned, normalized_output_after_correction=without_fence, **context)
    cleaned = without_fence
    repaired = repariere_json_strings(cleaned)
    if repaired != cleaned:
        _json_context_correction(context, "newline_repair_applied")
        record_json_pipeline_event(stage="newline_repair_applied", status="corrected", correction_type="newline_repair_applied", raw_output_before_correction=cleaned, normalized_output_after_correction=repaired, **context)
    cleaned = repaired
    decoder = json.JSONDecoder()
    try:
        parsed = json.loads(cleaned)
        record_json_pipeline_event(stage="json_parse_succeeded", status="valid", parsed_output_before_normalization=parsed, **context)
        return parsed
    except Exception as exc:
        record_json_pipeline_event(stage="json_parse_failed", status="retryable", validation_errors=[str(exc)], raw_output_before_correction=cleaned, **context)
        pass
    for match in re.finditer(r"\{", cleaned):
        try:
            parsed, _end = decoder.raw_decode(cleaned[match.start():])
            if isinstance(parsed, dict):
                _json_context_correction(context, "json_object_extracted_from_text")
                record_json_pipeline_event(stage="json_parse_succeeded", status="corrected", correction_type="json_object_extracted_from_text", raw_output_before_correction=cleaned, parsed_output_before_normalization=parsed, **context)
                return parsed
        except json.JSONDecodeError:
            continue
    match = re.search(r"\{.*\}", cleaned, re.DOTALL)
    if match:
        extracted = repariere_json_strings(match.group(0))
        try:
            parsed = json.loads(extracted)
            _json_context_correction(context, "json_object_extracted_by_regex")
            record_json_pipeline_event(stage="json_parse_succeeded", status="corrected", correction_type="json_object_extracted_by_regex", raw_output_before_correction=cleaned, parsed_output_before_normalization=parsed, **context)
            return parsed
        except Exception as exc:
            record_json_pipeline_event(stage="json_parse_failed", status="failed", validation_errors=[str(exc)], raw_output_before_correction=extracted, **context)
            raise
    record_json_pipeline_event(stage="json_parse_failed", status="failed", validation_errors=["Kein JSON gefunden"], raw_output_before_correction=cleaned, **context)
    raise ValueError("Kein JSON gefunden")


def repariere_json_strings(text: str) -> str:
    """Escaped echte Zeilenumbrueche innerhalb von JSON-Strings."""
    ergebnis = []
    in_string = False
    escape_next = False
    for zeichen in text:
        if escape_next:
            ergebnis.append(zeichen)
            escape_next = False
        elif zeichen == "\\":
            ergebnis.append(zeichen)
            escape_next = True
        elif zeichen == '"':
            in_string = not in_string
            ergebnis.append(zeichen)
        elif zeichen == "\n" and in_string:
            ergebnis.append("\\n")
        elif zeichen == "\r" and in_string:
            pass
        else:
            ergebnis.append(zeichen)
    return "".join(ergebnis)


def sicher_int(wert, min_val=0, max_val=3, fallback=0, feld=None, modell=None, headline=None, strict=False) -> int:
    """
    Wandelt einen Wert sicher in einen Integer um.
    Gibt fallback zurueck wenn der Wert kein Integer ist oder ausserhalb
    des erlaubten Bereichs liegt. Verhindert Abstuerze wenn ein LLM
    einen Text statt einer Zahl in ein Score-Feld schreibt.
    """
    try:
        ergebnis = int(float(str(wert).strip()))
        if strict and not min_val <= ergebnis <= max_val:
            raise ValueError(f"Score-Feld '{feld or 'unbekannt'}' liegt ausserhalb {min_val}-{max_val}: {wert!r}")
        return max(min_val, min(max_val, ergebnis))
    except (ValueError, TypeError) as exc:
        if feld and wert is not None:
            modell_text = modell or "unbekanntes Modell"
            headline_text = headline or "unbekannte Headline"
            aktion = "Fehler wird gemeldet" if strict else f"Fallback {fallback} verwendet"
            log_print(f"[WARN] Score-Feld '{feld}' fuer Modell {modell_text} / Headline {headline_text} enthaelt keinen validen Integer: {wert!r} -> {aktion}")
        if strict:
            raise ValueError(f"Score-Feld '{feld or 'unbekannt'}' enthaelt keinen validen Integer: {wert!r}") from exc
        return fallback


def _ist_int_im_bereich(wert, min_val=0, max_val=12) -> bool:
    try:
        zahl = int(float(str(wert).strip()))
        return min_val <= zahl <= max_val
    except (ValueError, TypeError):
        return False


def normalisiere_bewerter_json(daten: dict, prompt_typ: str, modell=None, headline=None) -> dict:
    """Normalisiert Bewerter-JSON defensiv nach dem Parsing."""
    return validate_and_normalize_bewerter_json(daten, prompt_typ, modell=modell, headline=headline)


def kategorie_zu_zahl(kat) -> int:
    """Wandelt Kategorien in Zahlen um.

    Eingabe: Low/Medium/High oder alte deutsche Werte als Fallback.
    Ausgabe: 0, 1 oder 2.
    Diese Zahlen braucht Kappa und die Kategorie-Reduktion.
    """
    s = str(kat or "").strip().lower()
    if s in {"hoch", "high"}:
        return 2
    if s in {"mittel", "medium"}:
        return 1
    if s in {"niedrig", "low"}:
        return 0
    raise ValueError(f"Unbekannte Kategorie: {kat!r}")


def zahl_zu_kategorie(n: int) -> str:
    """Wandelt 0/1/2 in englische Kategorien um.

    Eingabe: Zahl.
    Ausgabe: Low, Medium oder High.
    """
    return ["Low", "Medium", "High"][max(0, min(2, int(n)))]


def score_zu_kategorie(score: int) -> str:
    """Berechnet eine Kategorie aus einem Score.

    Eingabe: Score 0 bis 12.
    Ausgabe: Low, Medium oder High.
    Das ist der Fallback, falls ein Modell die Kategorie vergisst.
    """
    score = sicher_int(score, 0, 12, 0, feld="total_score")
    if score <= 3:
        return "Low"
    if score <= 7:
        return "Medium"
    return "High"


def kategorie_normalisieren(kat, score=None) -> str:
    """Normalisiert Kategorien auf Low/Medium/High.

    Eingabe: Kategorie und optional Score.
    Ausgabe: englische Kategorie.
    Falls die Kategorie fehlt, wird der Score genutzt.
    """
    if kat is None and score is not None:
        return score_zu_kategorie(score)
    return zahl_zu_kategorie(kategorie_zu_zahl(kat))


def binary_label_zu_zahl(value, fallback_category=None) -> int:
    """Wandelt Hyperpartisan-Labels in 0/1 um.

    Eingabe: binary_label oder Kategorie-Fallback.
    Ausgabe: 0 = non-hyperpartisan, 1 = hyperpartisan.
    Analyse 2.2 nutzt bevorzugt das JSON-Feld binary_label.
    """
    s = str(value or "").strip().lower()
    if s in {"1", "true", "yes", "hyperpartisan", "biased"}:
        return 1
    if s in {"0", "false", "no", "non-hyperpartisan", "nonhyperpartisan", "neutral", "unbiased"}:
        return 0
    return 0 if kategorie_zu_zahl(fallback_category) == 0 else 1


def lyu_label_zu_binaer(value) -> int:
    """Normalisiert Lyu-Ground-Truth-Labels.

    Eingabe: Zahl oder Text aus der CSV.
    Ausgabe: 0 oder 1.
    So funktionieren verschiedene CSV-Schreibweisen.
    """
    if isinstance(value, (int, float)):
        return 1 if value >= 0.5 else 0
    return 1 if str(value or "").strip().lower() in {"1", "true", "hyperpartisan", "biased", "yes"} else 0


def normalize_bias_result(parsed: dict, prompt_type: str, modell=None, headline=None) -> dict:
    """Vereinheitlicht Bewerter-JSON aus beiden Bias-Arten.

    Eingabe: geparstes JSON und Prompt-Typ.
    Ausgabe: Dict mit total_score, gesamt, category und kategorie; beide Kategorie-Werte sind englisch.
    Das komplette JSON bleibt fuer den Reformulierer erhalten.
    """
    return validate_and_normalize_bewerter_json(parsed, prompt_type, modell=modell, headline=headline)


def cell_value(value):
    """Bereitet Werte fuer UI-Tabellen und Exports vor."""
    if value is None:
        return ""
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    return value


def bias_export_values(result: dict | None, prompt_type: str) -> dict:
    """Liest alle standardisierten Bewerter-Felder fehlertolerant aus."""
    result = result or {}
    values = {}
    for field in BIAS_EXPORT_FIELDS[prompt_type]:
        if field == "category":
            values[field] = cell_value(kategorie_normalisieren(result.get("category") or result.get("kategorie"), result.get("total_score", result.get("gesamt"))) if (result.get("category") or result.get("kategorie") or result.get("total_score") is not None or result.get("gesamt") is not None) else "")
        elif field == "total_score":
            values[field] = cell_value(result.get("total_score", result.get("gesamt", "")))
        else:
            values[field] = cell_value(result.get(field, ""))
    return values


def bias_success_metadata(result: dict | None = None) -> dict:
    """Erzeugt technische Statusfelder fuer gueltige Bewerter-Exports."""
    result = result or {}
    return {
        "status": "ok",
        "error_message": "",
        "provider": "",
        "model": "",
        "retry_count": result.get("json_retry_count", 0),
        "json_status": result.get("json_status", "valid"),
        "json_warnings": cell_value(result.get("json_warnings", result.get("validation_warnings", []))),
        "correction_applied": result.get("correction_applied", False),
        "raw_output_available": result.get("raw_output_available", False),
    }


def bias_error_metadata(exc: Exception, modell_config: dict | None) -> dict:
    """Erzeugt technische Statusfelder fuer endgueltig ungueltige Bewerter-Outputs."""
    provider = (modell_config or {}).get("provider", "")
    model = (modell_config or {}).get("name") or (modell_config or {}).get("model_id") or ""
    if isinstance(exc, JsonOutputError):
        provider = exc.provider or provider
        model = exc.model or model
        return {
            "status": "validation_error",
            "error_message": sanitize_for_logging(str(exc)),
            "provider": provider,
            "model": model,
            "retry_count": exc.retry_count,
            "json_status": "failed",
            "json_warnings": "",
            "correction_applied": "",
            "raw_output_available": bool(exc.raw_output),
        }
    return {
        "status": "error",
        "error_message": sanitize_for_logging(str(exc)),
        "provider": provider,
        "model": model,
        "retry_count": "",
        "json_status": "failed",
        "json_warnings": "",
        "correction_applied": "",
        "raw_output_available": "",
    }


def add_bias_columns(row: dict, prefix: str, result: dict | None, prompt_type: str, suffix: str = "") -> None:
    """Fuegt Bewerter-Felder mit Prefix oder Suffix zu einer Tabellenzeile hinzu."""
    values = bias_export_values(result, prompt_type)
    field_names = {"category": "Kategorie", "total_score": "Score"}
    for field, value in values.items():
        display = field_names.get(field, field)
        if suffix:
            column = f"{display}{suffix}"
        else:
            column = f"{prefix} {display}" if prefix else display
        row[column] = value


def prefixed_bias_values(result: dict | None, prompt_type: str, suffix: str) -> dict:
    """Erzeugt CSV/Detail-Spalten mit technischem Feldnamen plus Suffix."""
    values = bias_export_values(result, prompt_type)
    renamed = {}
    for field, value in values.items():
        name = "score" if field == "total_score" else field
        renamed[f"{name}{suffix}"] = value
    return renamed


def json_technik_values(result: dict | None, prefix: str = "") -> dict:
    """Technische JSON-Metadaten ohne Raw-Outputs fuer Export-/Audit-Tabellen."""
    result = result or {}
    return {
        f"{prefix}json_status": result.get("json_status", "valid" if result else ""),
        f"{prefix}json_warnings": cell_value(result.get("json_warnings", result.get("validation_warnings", []))),
        f"{prefix}correction_applied": result.get("correction_applied", False if result else ""),
        f"{prefix}retry_count": result.get("retry_count", result.get("json_retry_count", 0 if result else "")),
        f"{prefix}raw_output_available": result.get("raw_output_available", False if result else ""),
    }


def rewriter_bias_analysis_payload(result: dict, prompt_type: str) -> dict:
    """Erzeugt den fachlichen bias_analysis-Payload fuer den Reformulierer ohne technische Zusatzfelder."""
    dimensions = BIAS_DIMENSIONS[prompt_type]
    payload = {dimension: result.get(dimension) for dimension in dimensions}
    payload["total_score"] = result.get("total_score")
    payload["category"] = kategorie_normalisieren(result.get("category"), result.get("total_score"))
    if prompt_type == "hyperpartisan":
        payload["binary_label"] = _binary_label_normalisieren(result.get("binary_label"))
    evidence = result.get("dimension_evidence") if isinstance(result.get("dimension_evidence"), dict) else {}
    payload["dimension_evidence"] = {dimension: str(evidence.get(dimension, result.get(f"{dimension}_evidence", ""))) for dimension in dimensions}
    payload["reasoning"] = str(result.get("reasoning", "")).strip()
    return payload


def reformulierer_values(reform_json: dict | None) -> dict:
    """Liest alle Reformulierer-Felder fehlertolerant aus."""
    reform_json = reform_json or {}
    return {field: cell_value(reform_json.get(field, "")) for field in REFORMULIERER_FIELDS}


def evaluate_headline(headline: str, modell_config: dict, prompt_type: str, status_callback=None) -> dict:
    """Bewertet eine Headline mit dem passenden Bewerter-Prompt.

    Eingabe: Headline, Modell-Konfiguration, Bias-Art und optionaler Status-Callback.
    Ausgabe: normalisiertes Bewerter-JSON.
    Wird in Tab 2 und in Modus B von Tab 3 genutzt.
    """
    return call_llm_for_json(
        modell_config=modell_config,
        system_prompt=prompt_for(prompt_type, "bewerter"),
        user_input=headline,
        prompt_type=prompt_type,
        output_schema=bewerter_json_schema(prompt_type),
        max_json_retries=2,
        status_callback=status_callback,
    )


def reformulate_headline(original: str, bias_analysis: dict, modell_config: dict, prompt_type: str, status_callback=None) -> dict:
    """Reformuliert eine Headline mit dem passenden Reformulierer-Prompt.

    Eingabe: Original, komplettes Bewerter-JSON, Modell, Bias-Art und optionaler Status-Callback.
    Ausgabe: Reformulierer-JSON mit neutralized_headline.
    Das Bewerter-JSON wird als bias_analysis uebergeben.
    """
    model_name = (modell_config or {}).get("name") or (modell_config or {}).get("model_id") or "Unbekanntes Modell"
    json_context = {"provider": (modell_config or {}).get("provider") or "unknown", "model": model_name or "unknown", "pipeline_type": "rewriter", "prompt_type": prompt_type, "input_id": _input_id_aus_text(original), "original_headline": original, "retry_count": 0}
    normalized_bias = ensure_normalized_bias_analysis(bias_analysis, prompt_type, modell=model_name, headline=original, json_context=json_context)
    clean_bias_analysis = rewriter_bias_analysis_payload(normalized_bias, prompt_type)
    payload = json.dumps({"original_headline": original, "bias_analysis": clean_bias_analysis}, ensure_ascii=False, indent=2)
    record_json_pipeline_event(stage="rewriter_input_built", status="valid", normalized_bias_analysis=clean_bias_analysis, final_output={"original_headline": original, "bias_analysis": clean_bias_analysis}, **json_context)
    return call_llm_for_rewriter_json(
        modell_config=modell_config,
        system_prompt=prompt_for(prompt_type, "reformulierer"),
        user_input=payload,
        output_schema=rewriter_model_output_schema(),
        max_json_retries=2,
        status_callback=status_callback,
        headline=original,
        prompt_type=prompt_type,
    )


# ============================================================
# ABSCHNITT 5: METRIKEN, EMBEDDINGS UND EXPORT-HILFEN
# ============================================================

def kappa_interpretation(k: float) -> str:
    """Interpretiert Cohen's Kappa nach Landis & Koch.

    Eingabe: Kappa-Wert.
    Ausgabe: deutscher Kurztext.
    Kappa misst Uebereinstimmung und zieht Zufallstreffer ab.
    """
    if k < 0.2:
        return "schlecht"
    if k < 0.4:
        return "fair"
    if k < 0.6:
        return "moderat"
    if k < 0.8:
        return "substanziell"
    return "fast perfekt"


def compute_kappa_table(model_categories: dict[str, list]) -> pd.DataFrame:
    """Berechnet Kappa fuer alle Modellpaare.

    Eingabe: pro Modell eine Liste von Kategorien als Zahlen.
    Ausgabe: Tabelle mit Modell A, Modell B, Kappa und Interpretation.
    Fehlende Werte werden paarweise uebersprungen.
    """
    rows = []
    for a, b in itertools.combinations(model_categories.keys(), 2):
        pairs = [(x, y) for x, y in zip(model_categories[a], model_categories[b]) if x is not None and y is not None]
        if len(pairs) < 2:
            rows.append({"Modell A": a, "Modell B": b, "Kappa": None, "Interpretation": "zu wenig Daten", "N": len(pairs)})
            continue
        k = float(cohen_kappa_score([p[0] for p in pairs], [p[1] for p in pairs]))
        rows.append({"Modell A": a, "Modell B": b, "Kappa": round(k, 3), "Interpretation": kappa_interpretation(k), "N": len(pairs)})
    return pd.DataFrame(rows)


def binary_metrics(y_true_by_model: dict, y_pred_by_model: dict, true_labels: tuple[str, str] | None = None) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Berechnet Precision, Recall, F1 und Confusion Matrices.

    Eingabe: y_true und y_pred pro Modell.
    Ausgabe: Metrik-Tabelle und vertikal verkettete 2x2-Matrizen.
    F1 verbindet Precision und Recall in einem Wert.
    """
    metric_rows = []
    cm_rows = []
    true_labels = true_labels or ("True: Low (0)", "True: Medium+High (1)")
    pred_labels = ("Pred: Low (0)", "Pred: Medium+High (1)")
    for model, y_true in y_true_by_model.items():
        y_pred = y_pred_by_model.get(model, [])
        if not y_true:
            metric_rows.append({"LLM": model, "Precision": None, "Recall": None, "F1": None, "N": 0})
            continue
        precision, recall, f1, _ = precision_recall_fscore_support(y_true, y_pred, average="binary", zero_division=0)
        metric_rows.append({"LLM": model, "Precision": round(float(precision), 3), "Recall": round(float(recall), 3), "F1": round(float(f1), 3), "N": len(y_true)})
        cm = confusion_matrix(y_true, y_pred, labels=[0, 1])
        cm_rows.extend([
            {"": f"Modell: {model}", pred_labels[0]: "", pred_labels[1]: ""},
            {"": true_labels[0], pred_labels[0]: int(cm[0][0]), pred_labels[1]: int(cm[0][1])},
            {"": true_labels[1], pred_labels[0]: int(cm[1][0]), pred_labels[1]: int(cm[1][1])},
            {"": "", pred_labels[0]: "", pred_labels[1]: ""},
        ])
    return pd.DataFrame(metric_rows), pd.DataFrame(cm_rows, columns=["", *pred_labels])


def category_reduced(before: str, after: str) -> bool:
    """Prueft, ob die Kategorie nachher niedriger ist.

    Eingabe: Kategorie vorher und nachher.
    Ausgabe: True oder False.
    Diese Logik wird in beiden Reformulierer-Analysen genutzt.
    """
    return kategorie_zu_zahl(after) < kategorie_zu_zahl(before)


def embed_modell():
    """Laedt das Embedding-Modell erst beim ersten Bedarf.

    Eingabe: keine.
    Ausgabe: SentenceTransformer-Modell.
    Lazy Load macht den App-Start schneller.
    """
    global _EMBED_MODELL
    if _EMBED_MODELL is None:
        from sentence_transformers import SentenceTransformer
        _EMBED_MODELL = SentenceTransformer("all-mpnet-base-v2")
    return _EMBED_MODELL


def cosine_similarity(text1: str, text2: str) -> float:
    """Berechnet semantische Aehnlichkeit zweier Texte.

    Eingabe: Original und Reformulierung.
    Ausgabe: Cosine Similarity.
    Hohe Werte bedeuten, dass der Inhalt wahrscheinlich erhalten blieb.
    """
    from sentence_transformers import util
    model = embed_modell()
    emb1 = model.encode(text1, convert_to_tensor=True)
    emb2 = model.encode(text2, convert_to_tensor=True)
    return float(util.cos_sim(emb1, emb2))


def csv_speichern(df: pd.DataFrame, name: str) -> str:
    """Speichert eine Tabelle als CSV.

    Eingabe: DataFrame und Analyse-Name.
    Ausgabe: Dateipfad.
    utf-8-sig sorgt dafuer, dass Excel Umlaute erkennt.
    """
    EXPORT_ORDNER.mkdir(exist_ok=True)
    path = EXPORT_ORDNER / f"ergebnisse_{name}_{timestamp()}.csv"
    df.to_csv(path, index=False, sep=CSV_SEP, encoding="utf-8-sig")
    return str(path)


def _is_dimension_score_column(column_name: str) -> bool:
    """Erkennt Dimensionsscore-Spalten fuer die Excel-Farbgebung."""
    normalized = str(column_name or "").strip().lower()
    dimensions = set(BIAS_DIMENSIONS["linguistic_bias"] + BIAS_DIMENSIONS["hyperpartisan"])
    return any(normalized == dim or normalized.endswith(f" {dim}") or normalized.endswith(f"_{dim}") for dim in dimensions) or normalized in {"score_vorher", "score_nachher"}


def _style_excel_workbook(writer) -> None:
    """Formatiert Excel-Sheets fuer Scores, Texte und Confusion Matrix."""
    from openpyxl.styles import Alignment, Font, PatternFill

    fill_yellow = PatternFill("solid", fgColor="FFFACD")
    fill_orange = PatternFill("solid", fgColor="FFD580")
    fill_red = PatternFill("solid", fgColor="FF6B6B")
    fill_green = PatternFill("solid", fgColor="90EE90")
    no_fill = PatternFill(fill_type=None)

    for worksheet in writer.book.worksheets:
        headers = [cell.value for cell in worksheet[1]]
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
        for column_index, header in enumerate(headers, start=1):
            header_text = str(header or "")
            header_lower = header_text.lower()
            wrap_column = any(token in header_lower for token in ["evidence", "reasoning", "neutralization_summary", "changed_terms"])
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=column_index)
                if column_index == 1 and isinstance(cell.value, str) and cell.value.startswith("Modell:"):
                    cell.font = Font(bold=True)
                if wrap_column:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                if _is_dimension_score_column(header_text):
                    try:
                        value = float(cell.value)
                    except (TypeError, ValueError):
                        continue
                    if value == 0:
                        cell.fill = no_fill
                    elif value == 1:
                        cell.fill = fill_yellow
                    elif value == 2:
                        cell.fill = fill_orange
                    elif value >= 3:
                        cell.fill = fill_red
                elif header_lower == "delta_score" or header_text == "Delta Score":
                    try:
                        value = float(cell.value)
                    except (TypeError, ValueError):
                        continue
                    if value < 0:
                        cell.fill = fill_green
                    elif value > 0:
                        cell.fill = fill_red
                    else:
                        cell.fill = no_fill
                elif header_lower == "changed_meaning_risk":
                    value = str(cell.value or "").strip().lower()
                    if value == "medium":
                        cell.fill = fill_yellow
                    elif value == "high":
                        cell.fill = fill_red
                    else:
                        cell.fill = no_fill
        for column_cells in worksheet.columns:
            values = [str(cell.value or "") for cell in column_cells[:50]]
            width = min(max((len(value) for value in values), default=10) + 2, 60)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width


def excel_speichern(sheets: dict[str, pd.DataFrame], name: str) -> str:
    """Speichert mehrere Tabellen in eine Excel-Datei.

    Eingabe: Sheetname -> DataFrame.
    Ausgabe: Dateipfad.
    Wird fuer Einzelanalysen und Gesamtexport genutzt.
    """
    EXPORT_ORDNER.mkdir(exist_ok=True)
    path = EXPORT_ORDNER / f"ergebnisse_{name}_{timestamp()}.xlsx"
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet, df in sheets.items():
            (df if df is not None else pd.DataFrame()).to_excel(writer, sheet_name=sheet[:31], index=False)
        _style_excel_workbook(writer)
    return str(path)


def export_bundle(primary_df: pd.DataFrame, sheets: dict[str, pd.DataFrame], name: str) -> list[str]:
    """Erzeugt CSV und XLSX fuer eine einzelne Analyse.

    Eingabe: Haupttabelle, Sheets und Analyse-Name.
    Ausgabe: Liste mit Dateipfaden.
    Bei leeren Daten wird nichts exportiert.
    """
    if primary_df is None or primary_df.empty:
        return []
    excel_sheets = {"Ergebnisse": primary_df, **sheets}
    return [csv_speichern(primary_df, name), excel_speichern(excel_sheets, name)]


# ============================================================
# ABSCHNITT 6: TAB 2 - BEWERTER-ANALYSEN
# ============================================================

def mini_tabelle_erstellen(headlines_text):
    """Erzeugt bis zu 50 Annotation-Zeilen.

    Eingabe: Textarea mit Headlines.
    Ausgabe: State plus UI-Updates fuer Markdown und Radio-Buttons.
    Analyse 2.1 nutzt diese manuelle Ground Truth.
    """
    headlines = [h.strip() for h in (headlines_text or "").splitlines() if h.strip()][:MAX_MINI]
    row_updates, md_updates, radio_updates = [], [], []
    for i in range(MAX_MINI):
        visible = i < len(headlines)
        row_updates.append(gr.update(visible=visible))
        md_updates.append(gr.update(value=f"**{i + 1}.** {headlines[i]}" if visible else ""))
        radio_updates.append(gr.update(value="Low"))
    return [headlines] + row_updates + md_updates + radio_updates


def labels_speichern(headlines, *labels):
    """Speichert manuelle Labels aus Analyse 2.1 im UI-State.

    Eingabe: Headline-Liste und Radio-Werte.
    Ausgabe: Annotationen-Dict und Status.
    Die Labels dienen als Ground Truth.
    """
    if not headlines:
        return {}, "Bitte zuerst Headlines eingeben und Annotationstabelle erstellen."
    annotations = {}
    for headline, label in zip(headlines, labels[:len(headlines)]):
        try:
            annotations[headline] = kategorie_normalisieren(label)
        except ValueError:
            annotations[headline] = "Low"
    return annotations, f"{len(annotations)} Labels gespeichert."


def analyse_2_1(annotations, selected_models):
    """Fuehrt Analyse 2.1 gegen eigene Ground Truth aus.

    Eingabe: manuelle Labels und LLM-Liste.
    Ausgabe: Vergleich, Kappa, Metriken, Confusion Matrix, Fehler, Status.
    Vollstaendige Bewerter-JSONs werden in ERGEBNISSE gespeichert.
    """
    empty = (pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame())
    no_rewriter_update = (gr.update(), gr.update())
    if not annotations:
        yield (*empty, None, *no_rewriter_update, "Bitte zuerst Labels speichern.")
        return
    if not selected_models:
        yield (*empty, None, *no_rewriter_update, "Bitte mindestens ein LLM auswaehlen.")
        return

    headlines = list(annotations.keys())
    raw_results, comparison_rows, export_rows, mistakes, errors = {}, [], [], [], []
    y_true_by_model = {m: [] for m in selected_models}
    y_pred_by_model = {m: [] for m in selected_models}
    categories_by_model = {m: [] for m in selected_models}

    for idx, headline in enumerate(headlines):
        user_label = annotations[headline]
        user_label = kategorie_normalisieren(user_label)
        user_bin = 0 if user_label == "Low" else 1
        row = {"Headline": headline, "Mein Label": user_label}
        all_ok = True
        for model_name in selected_models:
            yield (*empty, None, *no_rewriter_update, f"Bewerte Headline {idx + 1}/{len(headlines)} mit {model_name} ...")
            modell_config = config_finden(model_name)
            try:
                result = evaluate_headline(headline, modell_config, "linguistic_bias")
                raw_results[(idx, model_name)] = result
                kat, score = result["category"], result["total_score"]
                pred_bin = 0 if kat == "Low" else 1
                add_bias_columns(row, model_name, result, "linguistic_bias")
                row[f"{model_name} Übereinstimmung"] = "Ja" if pred_bin == user_bin else "Nein"
                export_row = {"headline": headline, "mein_label": user_label, "llm": model_name}
                export_row.update(bias_export_values(result, "linguistic_bias"))
                export_row.update(bias_success_metadata(result))
                export_row["uebereinstimmung"] = "Ja" if pred_bin == user_bin else "Nein"
                export_rows.append(export_row)
                categories_by_model[model_name].append(kategorie_zu_zahl(kat))
                y_true_by_model[model_name].append(user_bin)
                y_pred_by_model[model_name].append(pred_bin)
                if pred_bin != user_bin:
                    all_ok = False
                    mistakes.append({"LLM": model_name, "Headline": headline, "Mein Label": user_label, "LLM-Label": kat, "Score": score})
            except Exception as exc:
                add_bias_columns(row, model_name, None, "linguistic_bias")
                row[f"{model_name} Kategorie"] = "FEHLER"
                row[f"{model_name} Übereinstimmung"] = "-"
                export_row = {"headline": headline, "mein_label": user_label, "llm": model_name}
                export_row.update(bias_export_values(None, "linguistic_bias"))
                export_row["category"] = "FEHLER"
                export_row.update(bias_error_metadata(exc, modell_config))
                export_row["uebereinstimmung"] = "-"
                export_rows.append(export_row)
                categories_by_model[model_name].append(None)
                all_ok = False
                errors.append(f"{model_name} / Headline {idx + 1}: {exc}")
        row["Übereinstimmung"] = "Ja" if all_ok else "Nein"
        comparison_rows.append(row)

    df_comp = pd.DataFrame(comparison_rows)
    df_kappa = compute_kappa_table(categories_by_model) if len(selected_models) >= 2 else pd.DataFrame(columns=["Modell A", "Modell B", "Kappa", "Interpretation", "N"])
    df_metrics, df_cm = binary_metrics(y_true_by_model, y_pred_by_model)
    df_mistakes = pd.DataFrame(mistakes).groupby("LLM", group_keys=False).head(5) if mistakes else pd.DataFrame(columns=["LLM", "Headline", "Mein Label", "LLM-Label", "Score"])
    df_export = pd.DataFrame(export_rows, columns=["headline", "mein_label", "llm", *BIAS_EXPORT_FIELDS["linguistic_bias"], *BIAS_ERROR_FIELDS, "uebereinstimmung"])
    ERGEBNISSE["bewerter_2_1"] = {"headlines": headlines, "models": list(selected_models), "raw_results": raw_results, "user_labels": [annotations[h] for h in headlines], "comparison": df_comp, "export": df_export, "kappa": df_kappa, "metrics": df_metrics, "confusion": df_cm, "mistakes": df_mistakes}
    status = f"Fertig. {len(headlines)} Headlines x {len(selected_models)} LLMs."
    if errors:
        status += "\nFehler (Auszug): " + "; ".join(errors[:5])
    state = ERGEBNISSE["bewerter_2_1"]
    yield df_comp, df_kappa, df_metrics, df_cm, df_mistakes, state, *reformulierer_ui_updates_fuer_state(state, "bewerter_2_1"), status


def lyu_csv_vorschau(file):
    """Liest eine Lyu-CSV fuer die Vorschau.

    Eingabe: Gradio-Datei.
    Ausgabe: Vorschau, Spalten-Dropdowns und Status.
    title und label werden automatisch vorgeschlagen.
    """
    if file is None:
        return pd.DataFrame(), gr.update(choices=[], value=None), gr.update(choices=[], value=None), "Bitte CSV-Datei hochladen."
    path = file.name if hasattr(file, "name") else file
    try:
        df = pd.read_csv(path)
    except Exception as exc:
        return pd.DataFrame(), gr.update(choices=[], value=None), gr.update(choices=[], value=None), f"CSV-Fehler: {exc}"
    cols = list(df.columns)
    text_col = "title" if "title" in cols else (cols[0] if cols else None)
    label_col = "label" if "label" in cols else (cols[1] if len(cols) > 1 else None)
    return df.head(5), gr.update(choices=cols, value=text_col), gr.update(choices=cols, value=label_col), f"Geladen: {len(df)} Zeilen, {len(cols)} Spalten."


def analyse_2_2(file, text_col, label_col, n, selected_models):
    """Fuehrt externe Validierung mit dem Lyu-Datensatz aus.

    Eingabe: CSV, Spalten, Stichprobengroesse und LLMs.
    Ausgabe: Vergleich, Metriken, Confusion Matrix, Fehler, Status.
    Vollstaendige Hyperpartisan-JSONs werden gespeichert.
    """
    empty = (pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame())
    no_rewriter_update = (gr.update(), gr.update())
    if file is None:
        yield (*empty, None, *no_rewriter_update, "Bitte CSV-Datei hochladen.")
        return
    if not text_col or not label_col:
        yield (*empty, None, *no_rewriter_update, "Bitte Text- und Label-Spalte auswaehlen.")
        return
    if not selected_models:
        yield (*empty, None, *no_rewriter_update, "Bitte mindestens ein LLM auswaehlen.")
        return

    try:
        df = pd.read_csv(file.name if hasattr(file, "name") else file)
    except Exception as exc:
        yield (*empty, None, *no_rewriter_update, f"CSV-Fehler: {exc}")
        return
    sample = df.sample(n=min(int(n), len(df)), random_state=RANDOM_STATE).reset_index(drop=True)
    headlines = [str(x).strip() for x in sample[text_col].tolist()]
    external_labels = [lyu_label_zu_binaer(x) for x in sample[label_col].tolist()]
    raw_results, comparison_rows, export_rows, mistakes, errors = {}, [], [], [], []
    y_true_by_model = {m: [] for m in selected_models}
    y_pred_by_model = {m: [] for m in selected_models}

    for idx, headline in enumerate(headlines):
        true_bin = external_labels[idx]
        row = {"Headline": headline, "Lyu-Label": true_bin}
        all_ok = True
        for model_name in selected_models:
            yield (*empty, None, *no_rewriter_update, f"Bewerte Lyu-Headline {idx + 1}/{len(headlines)} mit {model_name} ...")
            modell_config = config_finden(model_name)
            try:
                result = evaluate_headline(headline, modell_config, "hyperpartisan")
                raw_results[(idx, model_name)] = result
                kat = result["category"]
                pred_bin = binary_label_zu_zahl(result.get("binary_label"), kat)
                add_bias_columns(row, model_name, result, "hyperpartisan")
                row[f"{model_name} Pred"] = pred_bin
                row[f"{model_name} Übereinstimmung"] = "Ja" if pred_bin == true_bin else "Nein"
                export_row = {"headline": headline, "lyu_label": true_bin, "llm": model_name}
                export_row.update(bias_export_values(result, "hyperpartisan"))
                export_row.update(bias_success_metadata(result))
                export_row["uebereinstimmung"] = "Ja" if pred_bin == true_bin else "Nein"
                export_rows.append(export_row)
                y_true_by_model[model_name].append(true_bin)
                y_pred_by_model[model_name].append(pred_bin)
                if pred_bin != true_bin:
                    all_ok = False
                    mistakes.append({"LLM": model_name, "Headline": headline, "Lyu-Label": true_bin, "LLM-Label": pred_bin, "Kategorie": kat, "Score": result["total_score"]})
            except Exception as exc:
                add_bias_columns(row, model_name, None, "hyperpartisan")
                row[f"{model_name} Kategorie"] = "FEHLER"
                row[f"{model_name} Pred"] = "FEHLER"
                row[f"{model_name} Übereinstimmung"] = "-"
                export_row = {"headline": headline, "lyu_label": true_bin, "llm": model_name}
                export_row.update(bias_export_values(None, "hyperpartisan"))
                export_row["category"] = "FEHLER"
                export_row.update(bias_error_metadata(exc, modell_config))
                export_row["uebereinstimmung"] = "-"
                export_rows.append(export_row)
                all_ok = False
                errors.append(f"{model_name} / Headline {idx + 1}: {exc}")
        row["Übereinstimmung"] = "Ja" if all_ok else "Nein"
        comparison_rows.append(row)

    df_comp = pd.DataFrame(comparison_rows)
    df_metrics, df_cm = binary_metrics(y_true_by_model, y_pred_by_model, ("True: Non-Hyperpartisan (0)", "True: Hyperpartisan (1)"))
    df_mistakes = pd.DataFrame(mistakes).groupby("LLM", group_keys=False).head(5) if mistakes else pd.DataFrame(columns=["LLM", "Headline", "Lyu-Label", "LLM-Label", "Kategorie", "Score"])
    df_export = pd.DataFrame(export_rows, columns=["headline", "lyu_label", "llm", *BIAS_EXPORT_FIELDS["hyperpartisan"], *BIAS_ERROR_FIELDS, "uebereinstimmung"])
    ERGEBNISSE["bewerter_2_2"] = {"headlines": headlines, "models": list(selected_models), "raw_results": raw_results, "external_labels": external_labels, "comparison": df_comp, "export": df_export, "metrics": df_metrics, "confusion": df_cm, "mistakes": df_mistakes, "sample": sample}
    status = f"Fertig. {len(headlines)} Headlines x {len(selected_models)} LLMs."
    if errors:
        status += "\nFehler (Auszug): " + "; ".join(errors[:5])
    state = ERGEBNISSE["bewerter_2_2"]
    yield df_comp, df_metrics, df_cm, df_mistakes, state, *reformulierer_ui_updates_fuer_state(state, "bewerter_2_2"), status


# ============================================================
# ABSCHNITT 7: TAB 3 - REFORMULIERER-ANALYSEN
# ============================================================

def state_models_for(key: str) -> list[str]:
    """Liest Modelle aus einem Bewerter-State.

    Eingabe: State-Schluessel.
    Ausgabe: Modellnamen mit vorhandenen Rohdaten.
    Modus A zeigt nur diese Modelle an.
    """
    state = ERGEBNISSE.get(key)
    return list(state.get("models", [])) if state else []


def update_mode_models(mode, state_key):
    """Aktualisiert LLM-Auswahl fuer Modus A oder B.

    Eingabe: Modus und State-Schluessel.
    Ausgabe: gr.update fuer CheckboxGroup.
    Modus A nutzt Tab-2-Modelle, Modus B alle konfigurierten.
    """
    if "übernehmen" in str(mode) or "uebernehmen" in str(mode):
        models = state_models_for(state_key)
        return gr.update(choices=models, value=models)
    names = llm_namen()
    return gr.update(choices=names, value=names)


def state_hat_bewerter_ergebnisse(state) -> bool:
    """Prueft, ob ein State echte Bewerter-Rohdaten fuer Reformulierer enthaelt."""
    return isinstance(state, dict) and bool(state.get("headlines")) and bool(state.get("raw_results"))


def state_ergebnis_anzahl(state) -> int:
    """Zaehlt vorhandene Bewerter-Ergebnisse im State."""
    return len(state.get("raw_results", {})) if isinstance(state, dict) else 0


def reformulierer_state_waehlen(explicit_state, fallback_key: str):
    """Nutzt expliziten Gradio-State und faellt nur bei Bedarf auf ERGEBNISSE zurueck."""
    if state_hat_bewerter_ergebnisse(explicit_state):
        return explicit_state
    fallback = ERGEBNISSE.get(fallback_key)
    if state_hat_bewerter_ergebnisse(fallback):
        return fallback
    return None


def threshold_preview(state_key: str, threshold, selected_models, only_external_positive=False):
    """Berechnet die Live-Vorschau fuer den Schwellenwert.

    Eingabe: State, Schwelle, Modelle und optional Lyu-Filter.
    Ausgabe: Text.
    Diese Vorschau verursacht keine LLM-Calls.
    """
    state = ERGEBNISSE.get(state_key)
    return threshold_preview_for_state(state, threshold, selected_models, only_external_positive)


def threshold_preview_for_state(state, threshold, selected_models, only_external_positive=False):
    """Berechnet die Schwellenwert-Vorschau aus einem expliziten State oder Fallback-State."""
    if not state:
        return "Keine Bewerter-Ergebnisse im State. Bitte zuerst Analyse 2 ausfuehren oder Modus B nutzen."
    selected_models = selected_models or state.get("models", [])
    threshold = int(threshold)
    count = 0
    for idx, _headline in enumerate(state.get("headlines", [])):
        if only_external_positive and state.get("external_labels", [])[idx] != 1:
            continue
        for model in selected_models:
            result = state.get("raw_results", {}).get((idx, model))
            if result and sicher_int(result.get("total_score", result.get("gesamt", 0)), 0, 12, 0, feld="total_score", modell=model, headline=_headline) >= threshold:
                count += 1
    calls = count * 2
    return f"Bei Schwelle {threshold}: {count} Headline-LLM-Paare = ca. {calls} LLM-Calls ({count} Reformulierungen + {count} Re-Bewertungen)."


def reformulierer_ui_updates_fuer_state(state, state_key: str, threshold=4, only_external_positive=False):
    """Aktualisiert Reformulierer-Modellauswahl und Vorschau nach erfolgreicher Bewerter-Analyse."""
    if state_hat_bewerter_ergebnisse(state):
        models = list(state.get("models", []))
        return gr.update(choices=models, value=models), threshold_preview_for_state(state, threshold, models, only_external_positive)
    return gr.update(), threshold_preview(state_key, threshold, state_models_for(state_key), only_external_positive)


def load_direct_linguistic(headlines_text, selected_models):
    """Bewertet neue Headlines fuer Reformulierer-Modus B.

    Eingabe: Textarea und LLMs.
    Ausgabe: State-artige Struktur und Fehlerliste.
    Dieser Modus braucht die Vorher-Bewertung neu.
    """
    headlines = [h.strip() for h in (headlines_text or "").splitlines() if h.strip()]
    raw_results, errors = {}, []
    for idx, headline in enumerate(headlines):
        for model in selected_models:
            try:
                raw_results[(idx, model)] = evaluate_headline(headline, config_finden(model), "linguistic_bias")
            except Exception as exc:
                errors.append(f"{model} / Headline {idx + 1}: {exc}")
    return {"headlines": headlines, "models": list(selected_models), "raw_results": raw_results}, errors


def load_direct_hyperpartisan(file, text_col, label_col, n, selected_models):
    """Bewertet eine neue CSV fuer Reformulierer-Modus B.

    Eingabe: CSV, Spalten, Stichprobengroesse und LLMs.
    Ausgabe: State-artige Struktur und Fehlerliste.
    Sampling nutzt random_state=42 fuer Reproduzierbarkeit.
    """
    if file is None:
        raise ValueError("Bitte CSV-Datei hochladen.")
    df = pd.read_csv(file.name if hasattr(file, "name") else file)
    sample = df.sample(n=min(int(n), len(df)), random_state=RANDOM_STATE).reset_index(drop=True)
    headlines = [str(x).strip() for x in sample[text_col].tolist()]
    external_labels = [lyu_label_zu_binaer(x) for x in sample[label_col].tolist()]
    raw_results, errors = {}, []
    for idx, headline in enumerate(headlines):
        for model in selected_models:
            try:
                raw_results[(idx, model)] = evaluate_headline(headline, config_finden(model), "hyperpartisan")
            except Exception as exc:
                errors.append(f"{model} / Headline {idx + 1}: {exc}")
    return {"headlines": headlines, "models": list(selected_models), "raw_results": raw_results, "external_labels": external_labels, "sample": sample}, errors


def run_reformulation_pipeline(state, selected_models, threshold, prompt_type, result_key, only_external_positive=False):
    """Fuehrt Reformulierung und Re-Bewertung aus.

    Eingabe: vorhandene Vorher-Bewertungen, Modelle, Schwelle und Bias-Art.
    Ausgabe: Detailtabelle, Zusammenfassung und Status.
    Nur Zeilen ueber der Schwelle erzeugen zwei LLM-Calls.
    """
    threshold = int(threshold)
    rows, errors = [], []
    headlines = state.get("headlines", [])
    selected_models = selected_models or state.get("models", [])

    def result_row(model, headline, before, reform_json, after, delta_score, reduced, similarity, status):
        row = {"headline": headline, "llm": model}
        row.update(prefixed_bias_values(before, prompt_type, "_vorher"))
        row.update(json_technik_values(before, "vorher_"))
        row.update(reformulierer_values(reform_json))
        row.update(json_technik_values(reform_json, "reformulierer_"))
        row.update(prefixed_bias_values(after, prompt_type, "_nachher"))
        row.update(json_technik_values(after, "nachher_"))
        row.update({
            "delta_score": delta_score,
            "kategorie_reduktion": reduced,
            "cosine_similarity": similarity,
            "status": status,
        })
        return row

    for idx, headline in enumerate(headlines):
        external_labels = state.get("external_labels", [])
        if only_external_positive and idx < len(external_labels) and external_labels[idx] != 1:
            for model in selected_models:
                before = state.get("raw_results", {}).get((idx, model))
                if before:
                    cfg = config_finden(model) or {}
                    json_context = {"provider": cfg.get("provider") or before.get("provider") or "unknown", "model": model or cfg.get("model_id") or "unknown", "pipeline_type": "rewriter", "prompt_type": prompt_type, "input_id": _input_id_aus_text(headline), "original_headline": headline, "retry_count": before.get("retry_count", before.get("json_retry_count", 0))}
                    before = ensure_normalized_bias_analysis(before, prompt_type, modell=model, headline=headline, json_context=json_context)
                    rows.append(result_row(model, headline, before, {"neutralized_headline": "Lyu-Label nicht hyperpartisan - nicht reformuliert"}, before, 0, "Nein", 1.0, "Lyu-Filter"))
            continue
        for model in selected_models:
            before = state.get("raw_results", {}).get((idx, model))
            if not before:
                continue
            cfg = config_finden(model) or {}
            json_context = {"provider": cfg.get("provider") or before.get("provider") or "unknown", "model": model or cfg.get("model_id") or "unknown", "pipeline_type": "rewriter", "prompt_type": prompt_type, "input_id": _input_id_aus_text(headline), "original_headline": headline, "retry_count": before.get("retry_count", before.get("json_retry_count", 0))}
            before = ensure_normalized_bias_analysis(before, prompt_type, modell=model, headline=headline, json_context=json_context)
            state.get("raw_results", {})[(idx, model)] = before
            score_before = sicher_int(before.get("total_score", before.get("gesamt", 0)), 0, 12, 0, feld="total_score", modell=model, headline=headline)
            kat_before = before.get("category") or before.get("kategorie")
            if score_before < threshold:
                rows.append(result_row(model, headline, before, {"neutralized_headline": "unter Schwelle - nicht reformuliert"}, before, 0, "Nein", 1.0, "unter Schwelle"))
                continue
            yield None, None, f"Reformuliere Headline {idx + 1}/{len(headlines)} mit {model} ..."
            try:
                cfg = config_finden(model)
                reform_json = reformulate_headline(headline, before, cfg, prompt_type)
                neutralized = str(reform_json.get("neutralized_headline", headline)).strip()
                yield None, None, f"Re-bewerte reformulierte Headline {idx + 1}/{len(headlines)} mit {model} ..."
                after = evaluate_headline(neutralized, cfg, prompt_type)
                sim = cosine_similarity(headline, neutralized)
                score_after = sicher_int(after.get("total_score", after.get("gesamt", 0)), 0, 12, 0, feld="total_score", modell=model, headline=neutralized)
                kat_after = after.get("category") or after.get("kategorie")
                rows.append(result_row(model, headline, before, reform_json, after, score_before - score_after, "Ja" if category_reduced(kat_before, kat_after) else "Nein", round(sim, 3), "reformuliert"))
            except Exception as exc:
                errors.append(f"{model} / Headline {idx + 1}: {exc}")
    df = pd.DataFrame(rows)
    summary_rows = []
    if not df.empty:
        for model, group in df.groupby("llm"):
            changed = group[group["status"] == "reformuliert"]
            base = changed if not changed.empty else group
            summary_rows.append({"LLM": model, "% Kategorie-Reduktion": round(base["kategorie_reduktion"].eq("Ja").mean() * 100, 1), "Ø Delta Score": round(float(base["delta_score"].mean()), 3), "Ø Cosine Similarity": round(float(base["cosine_similarity"].mean()), 3), "N": len(group)})
    summary = pd.DataFrame(summary_rows)
    df_export = df.drop(columns=["status"], errors="ignore")
    ERGEBNISSE[result_key] = {"detail": df, "export": df_export, "summary": summary, "threshold": threshold, "models": list(selected_models), "prompt_type": prompt_type}
    status = f"Fertig. {len(df)} Ergebniszeilen."
    if errors:
        status += "\nFehler (Auszug): " + "; ".join(errors[:5])
    yield df, summary, status


def analyse_3_1(mode, headlines_text, selected_models, threshold, state_2_1=None):
    """Fuehrt die Linguistic-Bias-Reformulierung aus.

    Eingabe: Modus, neue Headlines optional, LLMs und Schwelle.
    Ausgabe: Detailtabelle, Zusammenfassung und Status.
    Modus A nutzt Tab-2.1-Rohdaten ohne neue Vorher-Bewerter-Calls.
    """
    if "übernehmen" in str(mode) or "uebernehmen" in str(mode):
        state = reformulierer_state_waehlen(state_2_1, "bewerter_2_1")
        if not state:
            yield pd.DataFrame(), pd.DataFrame(), "Bitte zuerst Analyse 2.1 ausfuehren oder Modus B waehlen."
            return
        selected_models = selected_models or state.get("models", [])
        if not selected_models:
            yield pd.DataFrame(), pd.DataFrame(), "Bitte mindestens ein LLM auswaehlen."
            return
        log_print(f"[STATE] Analyse 3.1 nutzt {state_ergebnis_anzahl(state)} Bewerter-Ergebnisse aus Analyse 2.1.")
    else:
        if not selected_models:
            yield pd.DataFrame(), pd.DataFrame(), "Bitte mindestens ein LLM auswaehlen."
            return
        yield pd.DataFrame(), pd.DataFrame(), "Modus B: bewerte neue Headlines vor der Reformulierung ..."
        state, errors = load_direct_linguistic(headlines_text, selected_models)
        if errors:
            yield pd.DataFrame(), pd.DataFrame(), "Fehler bei Vorher-Bewertung: " + "; ".join(errors[:5])
            return
    for out in run_reformulation_pipeline(state, selected_models, threshold, "linguistic_bias", "reformulierer_3_1"):
        yield (pd.DataFrame(), pd.DataFrame(), out[2]) if out[0] is None else out


def analyse_3_2(mode, file, text_col, label_col, n, selected_models, threshold, only_positive, state_2_2=None):
    """Fuehrt die Hyperpartisan-Reformulierung aus.

    Eingabe: Modus, CSV optional, Spalten, LLMs, Schwelle und Lyu-Filter.
    Ausgabe: Detailtabelle, Zusammenfassung und Status.
    Modus A nutzt Tab-2.2-Rohdaten inklusive Stichprobe.
    """
    if "übernehmen" in str(mode) or "uebernehmen" in str(mode):
        state = reformulierer_state_waehlen(state_2_2, "bewerter_2_2")
        if not state:
            yield pd.DataFrame(), pd.DataFrame(), "Bitte zuerst Analyse 2.2 ausfuehren oder Modus B waehlen."
            return
        selected_models = selected_models or state.get("models", [])
        if not selected_models:
            yield pd.DataFrame(), pd.DataFrame(), "Bitte mindestens ein LLM auswaehlen."
            return
        log_print(f"[STATE] Analyse 3.2 nutzt {state_ergebnis_anzahl(state)} Bewerter-Ergebnisse aus Analyse 2.2.")
    else:
        if not selected_models:
            yield pd.DataFrame(), pd.DataFrame(), "Bitte mindestens ein LLM auswaehlen."
            return
        yield pd.DataFrame(), pd.DataFrame(), "Modus B: lade CSV und bewerte neue Stichprobe vor der Reformulierung ..."
        try:
            state, errors = load_direct_hyperpartisan(file, text_col, label_col, n, selected_models)
        except Exception as exc:
            yield pd.DataFrame(), pd.DataFrame(), str(exc)
            return
        if errors:
            yield pd.DataFrame(), pd.DataFrame(), "Fehler bei Vorher-Bewertung: " + "; ".join(errors[:5])
            return
    for out in run_reformulation_pipeline(state, selected_models, threshold, "hyperpartisan", "reformulierer_3_2", bool(only_positive)):
        yield (pd.DataFrame(), pd.DataFrame(), out[2]) if out[0] is None else out


def analyse_2_1_logged(*args):
    yield from analyse_mit_logging("analyse_2_1", analyse_2_1, *args)


def analyse_2_2_logged(*args):
    yield from analyse_mit_logging("analyse_2_2", analyse_2_2, *args)


def analyse_3_1_logged(*args):
    yield from analyse_mit_logging("analyse_3_1", analyse_3_1, *args)


def analyse_3_2_logged(*args):
    yield from analyse_mit_logging("analyse_3_2", analyse_3_2, *args)


# ============================================================
# ABSCHNITT 8: EXPORT UND DASHBOARD
# ============================================================

def export_2_1():
    """Exportiert Analyse 2.1 als CSV und XLSX.

    Eingabe: keine.
    Ausgabe: Dateiliste fuer Gradio.
    CSV ist die Vergleichstabelle, Excel enthaelt alle Tabellen.
    """
    s = ERGEBNISSE.get("bewerter_2_1")
    if not s:
        return []
    return export_bundle(s["export"], {"Vergleich": s["comparison"], "Kappa": s["kappa"], "Metriken": s["metrics"], "Confusion": s["confusion"], "Fehler": s["mistakes"]}, "bewerter_2_1")


def export_2_2():
    """Exportiert Analyse 2.2 als CSV und XLSX.

    Eingabe: keine.
    Ausgabe: Dateiliste fuer Gradio.
    Enthalten sind Vergleich, Metriken, Confusion Matrix und Fehler.
    """
    s = ERGEBNISSE.get("bewerter_2_2")
    if not s:
        return []
    return export_bundle(s["export"], {"Vergleich": s["comparison"], "Metriken": s["metrics"], "Confusion": s["confusion"], "Fehler": s["mistakes"]}, "bewerter_2_2")


def export_3_1():
    """Exportiert Analyse 3.1 als CSV und XLSX.

    Eingabe: keine.
    Ausgabe: Dateiliste fuer Gradio.
    Excel enthaelt Detail und Zusammenfassung.
    """
    s = ERGEBNISSE.get("reformulierer_3_1")
    if not s:
        return []
    return export_bundle(s["export"], {"Detail": s["detail"], "Zusammenfassung": s["summary"]}, "reformulierer_3_1")


def export_3_2():
    """Exportiert Analyse 3.2 als CSV und XLSX.

    Eingabe: keine.
    Ausgabe: Dateiliste fuer Gradio.
    Excel enthaelt Detail und Zusammenfassung.
    """
    s = ERGEBNISSE.get("reformulierer_3_2")
    if not s:
        return []
    return export_bundle(s["export"], {"Detail": s["detail"], "Zusammenfassung": s["summary"]}, "reformulierer_3_2")


def dashboard_refresh():
    """Sammelt die wichtigsten Tabellen fuer Tab 4.

    Eingabe: keine.
    Ausgabe: vier DataFrames.
    Das Dashboard zeigt den aktuellen Session-Stand.
    """
    b21 = ERGEBNISSE.get("bewerter_2_1") or {}
    b22 = ERGEBNISSE.get("bewerter_2_2") or {}
    r31 = ERGEBNISSE.get("reformulierer_3_1") or {}
    r32 = ERGEBNISSE.get("reformulierer_3_2") or {}
    return (
        b21.get("metrics", pd.DataFrame()),
        b22.get("metrics", pd.DataFrame()),
        r31.get("summary", pd.DataFrame()),
        r32.get("summary", pd.DataFrame()),
    )


def run_json_pipeline_selftest() -> None:
    """Fuehrt kleine lokale Tests fuer JSON-Parsing und Bewerter-Validierung aus."""
    valid = {
        "framing": 1,
        "intensifier": 2,
        "verb": 0,
        "labeling": 1,
        "dimension_evidence": {
            "framing": "uses \"loaded\" wording",
            "intensifier": "sharp",
            "verb": "says",
            "labeling": "controversial figure",
        },
        "reasoning": "Kurze Erklaerung mit Umlaut ae/oe/ue und Apostroph: don't.",
        "unexpected": "removed",
    }
    normalized = validate_and_normalize_bewerter_json(valid, "linguistic_bias")
    assert normalized["total_score"] == 4
    assert normalized["category"] == "Medium"
    assert normalized["kategorie"] == "Medium"
    assert "unexpected" not in normalized
    assert normalized["correction_applied"] is True

    clean_valid = {key: value for key, value in valid.items() if key != "unexpected"}
    clean_normalized = validate_and_normalize_bewerter_json(clean_valid, "linguistic_bias")
    assert clean_normalized["correction_applied"] is False
    assert clean_normalized["json_status"] == "valid"

    rewriter_payload = {"original_headline": "Test", "bias_analysis": rewriter_bias_analysis_payload(ensure_normalized_bias_analysis(clean_valid, "linguistic_bias"), "linguistic_bias")}
    assert rewriter_payload["bias_analysis"]["total_score"] == 4
    assert rewriter_payload["bias_analysis"]["category"] == "Medium"
    assert "json_status" not in rewriter_payload["bias_analysis"]
    assert "retry_count" not in rewriter_payload["bias_analysis"]
    assert "gesamt" not in rewriter_payload["bias_analysis"]
    assert "Niedrig" not in json.dumps(rewriter_payload, ensure_ascii=False)
    assert "Mittel" not in json.dumps(rewriter_payload, ensure_ascii=False)
    assert "Hoch" not in json.dumps(rewriter_payload, ensure_ascii=False)

    assert score_zu_kategorie(0) == "Low"
    assert score_zu_kategorie(3) == "Low"
    assert score_zu_kategorie(4) == "Medium"
    assert score_zu_kategorie(7) == "Medium"
    assert score_zu_kategorie(8) == "High"
    assert score_zu_kategorie(12) == "High"
    assert kategorie_normalisieren("Niedrig") == "Low"
    assert kategorie_normalisieren("Mittel") == "Medium"
    assert kategorie_normalisieren("Hoch") == "High"
    assert zahl_zu_kategorie(0) == "Low"
    assert zahl_zu_kategorie(1) == "Medium"
    assert zahl_zu_kategorie(2) == "High"

    with_wrong_derived = dict(valid, total_score=99, category="High")
    normalized = validate_and_normalize_bewerter_json(with_wrong_derived, "linguistic_bias")
    assert normalized["total_score"] == 4
    assert normalized["category"] == "Medium"
    assert normalized["correction_applied"] is True

    with_matching_derived = dict(clean_valid, total_score=4, category="Medium")
    normalized = validate_and_normalize_bewerter_json(with_matching_derived, "linguistic_bias")
    assert normalized["total_score"] == 4
    assert normalized["category"] == "Medium"
    assert normalized["correction_applied"] is True

    string_scores = dict(valid, framing="2", intensifier="0", verb="1", labeling="0")
    normalized = validate_and_normalize_bewerter_json(string_scores, "linguistic_bias")
    assert normalized["total_score"] == 3
    assert normalized["category"] == "Low"
    assert normalized["kategorie"] == "Low"
    assert normalized["correction_applied"] is True

    markdown = "Vorwort\n```json\n" + json.dumps(valid, ensure_ascii=False) + "\n```\nNachwort"
    normalized = validate_and_normalize_bewerter_json(json_aus_text(markdown), "linguistic_bias")
    assert normalized["framing"] == 1

    newline_json = '{"framing":1,"intensifier":0,"verb":0,"labeling":0,"dimension_evidence":{"framing":"first\nsecond","intensifier":"","verb":"","labeling":""},"reasoning":"ok"}'
    normalized = validate_and_normalize_bewerter_json(json_aus_text(newline_json), "linguistic_bias")
    assert "first\nsecond" == normalized["dimension_evidence"]["framing"]

    hyper_low = {
        "emotional_tone": 0,
        "one_sidedness": 1,
        "conflict_framing": 0,
        "identity_signaling": 0,
        "dimension_evidence": {"emotional_tone": "", "one_sidedness": "party label", "conflict_framing": "", "identity_signaling": ""},
        "reasoning": "Short explanation.",
    }
    normalized = validate_and_normalize_bewerter_json(hyper_low, "hyperpartisan")
    assert normalized["total_score"] == 1
    assert normalized["category"] == "Low"
    assert normalized["kategorie"] == "Low"
    assert normalized["binary_label"] == "non-hyperpartisan"
    rewriter_payload = {"original_headline": "Test", "bias_analysis": rewriter_bias_analysis_payload(ensure_normalized_bias_analysis(hyper_low, "hyperpartisan"), "hyperpartisan")}
    assert rewriter_payload["bias_analysis"]["binary_label"] == "non-hyperpartisan"
    assert rewriter_payload["bias_analysis"]["category"] == "Low"
    assert "json_status" not in rewriter_payload["bias_analysis"]
    assert "Niedrig" not in json.dumps(rewriter_payload, ensure_ascii=False)
    assert "Mittel" not in json.dumps(rewriter_payload, ensure_ascii=False)
    assert "Hoch" not in json.dumps(rewriter_payload, ensure_ascii=False)

    hyper_medium = dict(hyper_low, emotional_tone=2, one_sidedness=2, total_score=0, category="Low", binary_label="non-hyperpartisan")
    normalized = validate_and_normalize_bewerter_json(hyper_medium, "hyperpartisan")
    assert normalized["total_score"] == 4
    assert normalized["category"] == "Medium"
    assert normalized["binary_label"] == "hyperpartisan"

    hyper_high = dict(hyper_low, emotional_tone=3, one_sidedness=2, conflict_framing=2, identity_signaling=1, binary_label="neutral")
    normalized = validate_and_normalize_bewerter_json(hyper_high, "hyperpartisan")
    assert normalized["total_score"] == 8
    assert normalized["category"] == "High"
    assert normalized["binary_label"] == "hyperpartisan"

    export_values = bias_export_values(normalized, "hyperpartisan")
    assert export_values["category"] == "High"
    assert "Niedrig" not in json.dumps(export_values, ensure_ascii=False)
    assert "Mittel" not in json.dumps(export_values, ensure_ascii=False)
    assert "Hoch" not in json.dumps(export_values, ensure_ascii=False)

    invalid_cases = [
        ("missing evidence", {**valid, "dimension_evidence": {"framing": "x", "intensifier": "x", "verb": "x"}}),
        ("invalid score", {**valid, "framing": "safer place today"}),
        ("out of range score", {**valid, "framing": 4}),
        ("missing dimension", {key: value for key, value in valid.items() if key != "framing"}),
    ]
    for name, payload in invalid_cases:
        try:
            validate_and_normalize_bewerter_json(payload, "linguistic_bias")
            raise AssertionError(f"{name} wurde faelschlich akzeptiert")
        except ValueError:
            pass
    valid_rewriter = {
        "neutralized_headline": "Official says proposal changed",
        "changed_terms": [],
        "meaning_preservation": "The core claim is preserved.",
        "neutralization_summary": "Loaded wording was removed.",
        "changed_meaning_risk": "low",
    }
    normalized_rewriter = validate_and_normalize_rewriter_json(valid_rewriter)
    assert normalized_rewriter["changed_terms"] == []
    assert normalized_rewriter["changed_meaning_risk"] == "low"

    rewriter_with_terms = dict(valid_rewriter, neutralized_headline='Official says "proposal" changed', changed_terms=[{"dimension": "framing", "original": "war", "replacement": "dispute", "reason": "neutral wording"}], changed_meaning_risk="Medium")
    normalized_rewriter = validate_and_normalize_rewriter_json(rewriter_with_terms)
    assert normalized_rewriter["changed_meaning_risk"] == "medium"
    assert normalized_rewriter["changed_terms"][0]["dimension"] == "framing"

    markdown_rewriter = "```json\n" + json.dumps(valid_rewriter, ensure_ascii=False) + "\n```"
    assert validate_and_normalize_rewriter_json(json_aus_text(markdown_rewriter))["neutralized_headline"] == "Official says proposal changed"

    for name, payload in [
        ("rewriter score field", {**valid_rewriter, "total_score": 1}),
        ("rewriter reasoning field", {**valid_rewriter, "reasoning": "not allowed"}),
        ("rewriter bad changed_terms", {**valid_rewriter, "changed_terms": {}}),
    ]:
        try:
            validate_and_normalize_rewriter_json(payload)
            raise AssertionError(f"{name} wurde faelschlich akzeptiert")
        except ValueError:
            pass
    try:
        json_aus_text("komplett ungueltig")
        raise AssertionError("ungueltiger Rohtext wurde faelschlich geparst")
    except ValueError:
        pass
    print("JSON-Pipeline-Selftest erfolgreich.")


def _jsonl_events_lesen(path: Path) -> list[dict]:
    if not path.exists():
        return []
    return [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]


def run_json_pipeline_logging_selftest() -> None:
    """Testet strukturierte JSONL-Audit-Logs ohne externe API-Calls."""
    LOG_ORDNER.mkdir(exist_ok=True)
    run_id = f"selftest_json_pipeline_{datetime.now().strftime('%H%M%S_%f')}"
    jsonl_path = LOG_ORDNER / f"{run_id}_json_pipeline.jsonl"
    text_log_path = LOG_ORDNER / f"{run_id}.log"
    _AKTIVE_RUN_ID.set(run_id)
    _AKTIVE_LOG_DATEI.set(text_log_path)
    _AKTIVE_JSONL_LOG_DATEI.set(jsonl_path)
    _status_melden(None, "Rate Limit Testmeldung")

    valid_eval = {
        "emotional_tone": "2",
        "one_sidedness": 1,
        "conflict_framing": 0,
        "identity_signaling": 0,
        "total_score": 9,
        "category": "High",
        "binary_label": "non-hyperpartisan",
        "extra_comment": "remove me",
        "dimension_evidence": {"emotional_tone": "x", "one_sidedness": "y", "conflict_framing": "", "identity_signaling": ""},
        "reasoning": "ok",
    }
    raw_markdown = "```json\n" + json.dumps(valid_eval, ensure_ascii=False) + "\n```"
    parsed = json_aus_text(raw_markdown, json_context={"provider": "Test", "model": "mock", "pipeline_type": "evaluator", "prompt_type": "hyperpartisan", "input_id": "A", "original_headline": "Headline A", "retry_count": 0})
    normalized = validate_and_normalize_bewerter_json(parsed, "hyperpartisan", modell="mock", headline="Headline A", json_context={"provider": "Test", "model": "mock", "pipeline_type": "evaluator", "prompt_type": "hyperpartisan", "input_id": "A", "original_headline": "Headline A", "retry_count": 0})
    assert normalized["emotional_tone"] == 2
    assert normalized["total_score"] == 3
    assert normalized["category"] == "Low"
    assert normalized["binary_label"] == "non-hyperpartisan"
    assert normalized["json_status"] == "corrected"

    original_llm_bewerten = globals()["llm_bewerten"]
    try:
        responses = iter(["kein json", json.dumps(valid_eval, ensure_ascii=False)])
        globals()["llm_bewerten"] = lambda *args, **kwargs: next(responses)
        retry_result = call_llm_for_json(
            {"name": "mock", "provider": "Test", "model_id": "mock", "api_key": ""},
            "Return JSON",
            "Headline retry",
            "hyperpartisan",
            bewerter_json_schema("hyperpartisan"),
            max_json_retries=1,
        )
        assert retry_result["json_retry_count"] == 1
        assert retry_result["json_status"] == "retry_corrected"
        assert retry_result["correction_applied"] is True

        responses = iter(["kein json", "immer noch kein json"])
        globals()["llm_bewerten"] = lambda *args, **kwargs: next(responses)
        try:
            call_llm_for_json({"name": "mock", "provider": "Test", "model_id": "mock", "api_key": ""}, "Return JSON", "Headline fail", "hyperpartisan", bewerter_json_schema("hyperpartisan"), max_json_retries=1)
            raise AssertionError("ungueltiges JSON wurde nach Retry faelschlich akzeptiert")
        except JsonOutputError:
            pass
    finally:
        globals()["llm_bewerten"] = original_llm_bewerten

    valid_rewriter = {
        "neutralized_headline": "Neutral headline",
        "changed_terms": [],
        "meaning_preservation": "Meaning preserved.",
        "neutralization_summary": "Tone reduced.",
        "changed_meaning_risk": "Medium",
    }
    rewriter_result = validate_and_normalize_rewriter_json(valid_rewriter, modell="mock", headline="Headline R", json_context={"provider": "Test", "model": "mock", "pipeline_type": "rewriter", "prompt_type": "hyperpartisan", "input_id": "R", "original_headline": "Headline R", "retry_count": 0})
    assert rewriter_result["changed_meaning_risk"] == "medium"
    original_llm_bewerten = globals()["llm_bewerten"]
    try:
        responses = iter(["kein json", json.dumps(valid_rewriter, ensure_ascii=False)])
        globals()["llm_bewerten"] = lambda *args, **kwargs: next(responses)
        retry_rewriter = call_llm_for_rewriter_json(
            {"name": "mock", "provider": "Test", "model_id": "mock", "api_key": ""},
            "Return JSON",
            json.dumps({"original_headline": "Headline R", "bias_analysis": normalized}, ensure_ascii=False),
            rewriter_model_output_schema(),
            max_json_retries=1,
            headline="Headline R",
            prompt_type="hyperpartisan",
        )
        assert retry_rewriter["retry_count"] == 1
        assert retry_rewriter["json_status"] == "retry_corrected"
        assert retry_rewriter["correction_applied"] is True

        responses = iter(["kein json", "immer noch kein json"])
        globals()["llm_bewerten"] = lambda *args, **kwargs: next(responses)
        try:
            call_llm_for_rewriter_json({"name": "mock", "provider": "Test", "model_id": "mock", "api_key": ""}, "Return JSON", "{}", rewriter_model_output_schema(), max_json_retries=1, headline="Headline R", prompt_type="hyperpartisan")
            raise AssertionError("ungueltiges Reformulierer-JSON wurde nach Retry faelschlich akzeptiert")
        except JsonOutputError:
            pass
    finally:
        globals()["llm_bewerten"] = original_llm_bewerten
    try:
        validate_and_normalize_rewriter_json({**valid_rewriter, "total_score": 3}, modell="mock", headline="Headline R", json_context={"provider": "Test", "model": "mock", "pipeline_type": "rewriter", "prompt_type": "hyperpartisan", "input_id": "R2", "original_headline": "Headline R", "retry_count": 0})
        raise AssertionError("Reformulierer-Score-Feld wurde faelschlich akzeptiert")
    except ValueError:
        pass
    record_json_pipeline_event(stage="rewriter_input_built", status="valid", provider="Test", model="mock", pipeline_type="rewriter", prompt_type="hyperpartisan", input_id="R3", original_headline="Headline R", normalized_bias_analysis=normalized)

    events = _jsonl_events_lesen(jsonl_path)
    stages = [event.get("stage") for event in events]
    corrections = [event.get("correction_type") for event in events]
    assert all(event.get("audit_schema_version") == JSON_PIPELINE_AUDIT_SCHEMA_VERSION for event in events)
    assert all(event.get("provider") for event in events)
    assert all(event.get("model") for event in events)
    assert "raw_model_output_received" in stages
    assert "markdown_fence_removed" in stages
    assert "final_output_valid" in stages
    assert "retry_started" in stages
    assert "retry_succeeded" in stages
    assert "retry_failed" in stages
    assert "final_output_invalid" in stages
    assert "rewriter_input_built" in stages
    assert "score_type_normalized" in json.dumps(events, ensure_ascii=False)
    assert "derived_field_overwritten" in json.dumps(events, ensure_ascii=False)
    assert "unexpected_model_derived_fields" in corrections
    assert "rewriter_score_fields_detected" in corrections
    text_log = text_log_path.read_text(encoding="utf-8")
    assert "[JSON-REPAIR] JSON-Reparatur fuer Test/mock" in text_log
    assert "[RATE LIMIT] Rate Limit Testmeldung" in text_log
    assert "[RATE LIMIT] JSON-Reparatur" not in text_log
    print(f"JSON-Pipeline-Logging-Selftest erfolgreich: {jsonl_path}")


def run_json_pipeline_audit_security_selftest() -> None:
    """Testet Audit-Versionierung, Raw-Output-Datensparsamkeit und Secret-Redaction."""
    LOG_ORDNER.mkdir(exist_ok=True)
    run_id = f"selftest_json_security_{datetime.now().strftime('%H%M%S_%f')}"
    jsonl_path = LOG_ORDNER / f"{run_id}_json_pipeline.jsonl"
    text_log_path = LOG_ORDNER / f"{run_id}.log"
    _AKTIVE_RUN_ID.set(run_id)
    _AKTIVE_LOG_DATEI.set(text_log_path)
    _AKTIVE_JSONL_LOG_DATEI.set(jsonl_path)

    dummy_openai_key = "sk-proj-abcdefghijklmnopqrstuvwxyz1234567890"
    dummy_anthropic_key = "sk-ant-api03-abcdefghijklmnopqrstuvwxyz1234567890"
    dummy_google_key = "AIzaSyabcdefghijklmnopqrstuvwxyz1234567890"
    dummy_together_key = "tgp_v1_abcdefghijklmnopqrstuvwxyz1234567890"
    dummy_bearer = "Bearer abcdefghijklmnopqrstuvwxyz1234567890"
    raw_text = f"```json\n{{\"api_key\": \"{dummy_openai_key}\", \"text\": \"hello\"}}\n```"

    old_raw_setting = globals()["LOG_RAW_LLM_OUTPUTS"]
    old_raw_limit = globals()["LOG_RAW_LLM_OUTPUT_LIMIT"]
    try:
        globals()["LOG_RAW_LLM_OUTPUTS"] = False
        globals()["LOG_RAW_LLM_OUTPUT_LIMIT"] = 20
        record_json_pipeline_event(
            stage="raw_model_output_received",
            status="received",
            provider="Test",
            model="mock",
            pipeline_type="evaluator",
            prompt_type="hyperpartisan",
            input_id="S1",
            original_headline="Security headline",
            raw_output_before_correction=raw_text,
            nested={"api_key": dummy_openai_key, "headers": {"Authorization": dummy_bearer}, "items": [{"token": dummy_together_key}, "normal"]},
        )
        record_json_pipeline_event(
            stage="raw_truncation_short",
            status="received",
            provider="Test",
            model="mock",
            pipeline_type="evaluator",
            prompt_type="hyperpartisan",
            input_id="S_SHORT",
            raw_output_before_correction="short",
        )
        record_json_pipeline_event(
            stage="raw_truncation_exact",
            status="received",
            provider="Test",
            model="mock",
            pipeline_type="evaluator",
            prompt_type="hyperpartisan",
            input_id="S_EXACT",
            raw_output_before_correction="x" * 20,
        )
        record_json_pipeline_event(
            stage="raw_truncation_long",
            status="received",
            provider="Test",
            model="mock",
            pipeline_type="evaluator",
            prompt_type="hyperpartisan",
            input_id="S_LONG",
            raw_output_before_correction="x" * 21,
        )
        record_json_pipeline_event(
            stage="non_serializable_test",
            status="valid",
            provider="Test",
            model="mock",
            pipeline_type="evaluator",
            prompt_type="hyperpartisan",
            input_id="S2",
            odd_object=object(),
            raw_output_last_attempt=f"secret {dummy_anthropic_key} {dummy_google_key}",
        )
        globals()["LOG_RAW_LLM_OUTPUTS"] = True
        record_json_pipeline_event(
            stage="raw_model_output_received",
            status="received",
            provider="Test",
            model="mock",
            pipeline_type="evaluator",
            prompt_type="hyperpartisan",
            input_id="S3",
            raw_output_before_correction="full raw output for debug",
        )
        log_print(f"Authorization: {dummy_bearer} key={dummy_openai_key}")
    finally:
        globals()["LOG_RAW_LLM_OUTPUTS"] = old_raw_setting
        globals()["LOG_RAW_LLM_OUTPUT_LIMIT"] = old_raw_limit

    events = _jsonl_events_lesen(jsonl_path)
    assert events
    assert all(event.get("audit_schema_version") == JSON_PIPELINE_AUDIT_SCHEMA_VERSION for event in events)
    default_event = next(event for event in events if event.get("input_id") == "S1")
    assert "raw_output_before_correction" not in default_event
    assert "raw_output_before_correction_preview" in default_event
    assert default_event["raw_output_before_correction_truncated"] is True
    assert default_event["raw_output_before_correction_available"] is True
    assert default_event["raw_output_before_correction_length_chars"] == len(raw_text)
    short_event = next(event for event in events if event.get("input_id") == "S_SHORT")
    exact_event = next(event for event in events if event.get("input_id") == "S_EXACT")
    long_event = next(event for event in events if event.get("input_id") == "S_LONG")
    assert short_event["raw_output_before_correction_preview"] == "short"
    assert short_event["raw_output_before_correction_truncated"] is False
    assert exact_event["raw_output_before_correction_preview"] == "x" * 20
    assert exact_event["raw_output_before_correction_truncated"] is False
    assert long_event["raw_output_before_correction_preview"] == "x" * 20
    assert long_event["raw_output_before_correction_truncated"] is True
    full_event = next(event for event in events if event.get("input_id") == "S3")
    assert full_event["raw_output_before_correction"] == "full raw output for debug"
    assert full_event["raw_output_before_correction_truncated"] is False
    serialized = json.dumps(events, ensure_ascii=False)
    for secret in [dummy_openai_key, dummy_anthropic_key, dummy_google_key, dummy_together_key, dummy_bearer.split()[-1]]:
        assert secret not in serialized
        assert secret not in text_log_path.read_text(encoding="utf-8")
    assert "normal" in serialized

    error_meta = bias_error_metadata(JsonOutputError(f"kaputt {dummy_openai_key}", raw_output="voller raw output", provider="Test", model="mock", retry_count=2), {"provider": "Test", "name": "mock"})
    assert "raw_output" not in error_meta
    assert dummy_openai_key not in error_meta["error_message"]
    assert error_meta["raw_output_available"] is True
    df_export = pd.DataFrame([error_meta], columns=BIAS_ERROR_FIELDS)
    assert "raw_output" not in df_export.columns
    print(f"JSON-Pipeline-Audit-Security-Selftest erfolgreich: {jsonl_path}")


def run_state_and_export_selftest() -> None:
    """Testet Tab-2-zu-Tab-3-State-Uebergabe und Excel-Technikspalten ohne API-Calls."""
    original_evaluate = globals()["evaluate_headline"]
    original_reformulate = globals()["reformulate_headline"]
    original_similarity = globals()["cosine_similarity"]
    old_results = dict(ERGEBNISSE)

    def fake_evaluate(headline, modell_config, prompt_type, status_callback=None):
        if prompt_type == "linguistic_bias":
            data = {
                "framing": 1,
                "intensifier": 0,
                "verb": 0,
                "labeling": 0,
                "dimension_evidence": {"framing": "x", "intensifier": "", "verb": "", "labeling": ""},
                "reasoning": "ok",
            }
        else:
            data = {
                "emotional_tone": 1,
                "one_sidedness": 0,
                "conflict_framing": 0,
                "identity_signaling": 0,
                "dimension_evidence": {"emotional_tone": "x", "one_sidedness": "", "conflict_framing": "", "identity_signaling": ""},
                "reasoning": "ok",
            }
        result = validate_and_normalize_bewerter_json(data, prompt_type, modell="mock", headline=headline, json_context={"provider": "Test", "model": "mock", "pipeline_type": "evaluator", "prompt_type": prompt_type, "input_id": _input_id_aus_text(headline), "original_headline": headline, "retry_count": 0})
        result["retry_count"] = 0
        result["json_retry_count"] = 0
        return result

    def fake_reformulate(original, bias_analysis, modell_config, prompt_type, status_callback=None):
        return {
            "neutralized_headline": original,
            "changed_terms": [],
            "meaning_preservation": "preserved",
            "neutralization_summary": "none",
            "changed_meaning_risk": "low",
            "json_status": "valid",
            "json_warnings": [],
            "correction_applied": False,
            "retry_count": 0,
            "raw_output_available": True,
        }

    try:
        globals()["evaluate_headline"] = fake_evaluate
        globals()["reformulate_headline"] = fake_reformulate
        globals()["cosine_similarity"] = lambda *_args, **_kwargs: 1.0
        ERGEBNISSE.update({"bewerter_2_1": None, "bewerter_2_2": None, "reformulierer_3_1": None, "reformulierer_3_2": None})

        out_21 = list(analyse_2_1_logged({"Headline A": "Low"}, ["mock"]))[-1]
        state_21 = out_21[-4]
        assert state_hat_bewerter_ergebnisse(state_21)
        assert "Headline-LLM-Paare" in out_21[-2]
        assert "ca. 0 LLM-Calls" in out_21[-2]
        assert "Fertig. 1 Headlines" in out_21[-1]
        out_31 = list(analyse_3_1_logged("Ergebnisse aus Analyse 2.1 übernehmen", "", ["mock"], 4, state_21))[-1]
        assert "Bitte zuerst Analyse 2.1" not in out_31[-1]
        assert not out_31[0].empty
        out_31_empty_selection = list(analyse_3_1("Ergebnisse aus Analyse 2.1 übernehmen", "", [], 4, state_21))[-1]
        assert "Bitte mindestens ein LLM" not in out_31_empty_selection[-1]
        assert not out_31_empty_selection[0].empty

        ERGEBNISSE["bewerter_2_1"] = None
        no_state_31 = list(analyse_3_1("Ergebnisse aus Analyse 2.1 übernehmen", "", ["mock"], 4, None))[-1]
        assert "Bitte zuerst Analyse 2.1" in no_state_31[-1]
        mode_b_31 = list(analyse_3_1_logged("Neue Headlines eingeben und direkt bewerten", "Headline B", ["mock"], 4, None))[-1]
        assert "Bitte zuerst Analyse" not in mode_b_31[-1]
        assert not mode_b_31[0].empty

        csv_path = EXPORT_ORDNER / f"selftest_state_{datetime.now().strftime('%H%M%S_%f')}.csv"
        pd.DataFrame({"title": ["Hyper headline"], "label": [1]}).to_csv(csv_path, index=False)
        out_22 = list(analyse_2_2_logged(str(csv_path), "title", "label", 1, ["mock"]))[-1]
        state_22 = out_22[-4]
        assert state_hat_bewerter_ergebnisse(state_22)
        assert "Headline-LLM-Paare" in out_22[-2]
        assert "ca. 0 LLM-Calls" in out_22[-2]
        assert "Fertig. 1 Headlines" in out_22[-1]
        out_32 = list(analyse_3_2_logged("Ergebnisse aus Analyse 2.2 übernehmen", None, None, None, 1, ["mock"], 4, False, state_22))[-1]
        assert "Bitte zuerst Analyse 2.2" not in out_32[-1]
        assert not out_32[0].empty
        out_32_empty_selection = list(analyse_3_2("Ergebnisse aus Analyse 2.2 übernehmen", None, None, None, 1, [], 4, False, state_22))[-1]
        assert "Bitte mindestens ein LLM" not in out_32_empty_selection[-1]
        assert not out_32_empty_selection[0].empty
        ERGEBNISSE["bewerter_2_2"] = None
        no_state_32 = list(analyse_3_2("Ergebnisse aus Analyse 2.2 übernehmen", None, None, None, 1, ["mock"], 4, False, None))[-1]
        assert "Bitte zuerst Analyse 2.2" in no_state_32[-1]

        ERGEBNISSE["bewerter_2_1"] = state_21
        exported = export_2_1()
        xlsx_files = [path for path in exported if str(path).endswith(".xlsx")]
        assert xlsx_files
        exported_df = pd.read_excel(xlsx_files[0], sheet_name="Ergebnisse")
        for column in ["json_status", "json_warnings", "correction_applied", "retry_count", "raw_output_available"]:
            assert column in exported_df.columns
        assert "raw_output" not in exported_df.columns
    finally:
        globals()["evaluate_headline"] = original_evaluate
        globals()["reformulate_headline"] = original_reformulate
        globals()["cosine_similarity"] = original_similarity
        ERGEBNISSE.clear()
        ERGEBNISSE.update(old_results)
    print("State-und-Export-Selftest erfolgreich.")


def method_markdown() -> str:
    """Erzeugt die methodische Export-Erklaerung.

    Eingabe: keine.
    Ausgabe: Markdown-Text.
    Dokumentiert Metriken, Mappings und Quellen.
    """
    return """# Methodische Berechnungserklaerung

## Bias-Arten und Prompts
Die App verwendet zwei Bias-Arten: Linguistic Bias und Hyperpartisanship. Jede Bias-Art hat einen eigenen Bewerter-Prompt und Reformulierer-Prompt im Ordner `prompts/`.

## Kategorien und Scores
Alle Bewerter liefern einen Score von 0 bis 12. Intern werden Kategorien englisch normalisiert: 0-3 = Low, 4-7 = Medium, 8-12 = High.

## Cohen's Kappa
Cohen's Kappa misst die Uebereinstimmung zwischen zwei Bewertern und korrigiert um zufaellige Uebereinstimmung. Die Berechnung erfolgt auf Kategorieebene Low/Medium/High mit `sklearn.metrics.cohen_kappa_score`. Interpretation nach Landis & Koch (1977): <0.20 schlecht, 0.20-0.40 fair, 0.40-0.60 moderat, 0.60-0.80 substanziell, >0.80 fast perfekt.

## Precision, Recall und F1
Fuer Detection-Metriken wird binaer gemappt: Low = 0, Medium/High = 1. Bei Hyperpartisanship wird bevorzugt das Feld `binary_label` aus dem JSON verwendet; falls es fehlt, wird auf die Kategorie zurueckgegriffen.

## Reformulierer-Pipeline
Modus A uebernimmt vorhandene Bewerter-Rohdaten aus Tab 2. Pro reformulierter Headline und LLM entstehen dann zwei LLM-Calls: Reformulierung und Re-Bewertung. Modus B erzeugt zuerst neue Bewerter-Ergebnisse und braucht deshalb drei LLM-Calls pro Headline und LLM.

## Cosine Similarity
Die semantische Aehnlichkeit zwischen Original und Reformulierung wird mit `sentence-transformers/all-mpnet-base-v2` berechnet. Das Modell wird lazy geladen.

## Sampling
Alle Stichproben aus CSV-Dateien verwenden `random_state=42`.

## Quellen
Lyu et al. (2024), Menzner & Leidner (2024), Raza et al. (2024), Landis & Koch (1977), Recasens et al. (2013), Hamborg et al. (2019).
"""


def export_all():
    """Exportiert alle vorhandenen Ergebnisse.

    Eingabe: keine.
    Ausgabe: Dateiliste und Status.
    Es entstehen Einzel-CSVs, eine Gesamt-XLSX und eine Markdown-Erklaerung.
    """
    files = []
    sheets = {"Uebersicht": pd.DataFrame([{"App-Version": APP_VERSION, "Export": timestamp()}])}
    for key, label in [("bewerter_2_1", "Bewerter_2_1"), ("bewerter_2_2", "Bewerter_2_2"), ("reformulierer_3_1", "Reformulierer_3_1"), ("reformulierer_3_2", "Reformulierer_3_2")]:
        state = ERGEBNISSE.get(key)
        if not state:
            continue
        df = state.get("export") if isinstance(state.get("export"), pd.DataFrame) else (state.get("comparison") if isinstance(state.get("comparison"), pd.DataFrame) else state.get("detail"))
        if isinstance(df, pd.DataFrame) and not df.empty:
            files.append(csv_speichern(df, key))
            sheets[label] = df
        if isinstance(state.get("metrics"), pd.DataFrame):
            sheets[f"{label}_Metriken"] = state["metrics"]
        if isinstance(state.get("summary"), pd.DataFrame):
            sheets[f"{label}_Summary"] = state["summary"]
    mistakes = []
    for key in ["bewerter_2_1", "bewerter_2_2"]:
        state = ERGEBNISSE.get(key)
        if state and isinstance(state.get("mistakes"), pd.DataFrame):
            tmp = state["mistakes"].copy()
            tmp.insert(0, "Analyse", key)
            mistakes.append(tmp)
    if mistakes:
        sheets["Fehlklassifikationen"] = pd.concat(mistakes, ignore_index=True)
    xlsx = excel_speichern(sheets, "alle_ergebnisse")
    md_path = EXPORT_ORDNER / f"ergebnisse_methodik_{timestamp()}.md"
    md_path.write_text(method_markdown(), encoding="utf-8")
    files.extend([xlsx, str(md_path)])
    return files, f"Export fertig: {len(files)} Dateien erzeugt."


# ============================================================
# ABSCHNITT 9: GRADIO-UI
# ============================================================

ensure_basic_files()
migrate_prompt_files()

with gr.Blocks(title="BiasScore Evaluation") as app:
    gr.Markdown("# BiasScore Evaluation App")

    with gr.Tabs():
        with gr.Tab("1 - LLM-Konfiguration"):
            gr.Markdown("LLMs konfigurieren. API-Keys werden in `secrets.json` gespeichert und nicht in `config.json`.")
            with gr.Row():
                tf_name = gr.Textbox(label="Name", placeholder="z. B. GPT-4o Mini")
                dd_provider = gr.Dropdown(PROVIDER, value="OpenAI", label="Provider")
                tf_model_id = gr.Textbox(label="Modell-ID", placeholder="z. B. gpt-4o-mini")
            with gr.Row():
                tf_api_key = gr.Textbox(label="API-Key", type="password")
                tf_base_url = gr.Textbox(label="Basis-URL", placeholder="leer lassen fuer Gemini/Groq/Ollama-Standard")
            btn_save = gr.Button("LLM speichern", variant="primary")
            cfg_status = gr.Markdown("")
            llm_table = gr.Dataframe(value=tabelle_aus_config(), interactive=False, label="Gespeicherte LLMs")
            with gr.Row():
                dd_remove = gr.Dropdown(choices=llm_namen(), label="LLM entfernen")
                btn_remove = gr.Button("LLM entfernen")
            gr.Markdown("""
### Groq
API-Key unter https://console.groq.com erstellen. Provider `Groq` waehlen, Modell-ID z. B. `llama-3.1-8b-instant`, `llama-3.3-70b-versatile`, `mixtral-8x7b-32768`, `gemma2-9b-it` oder `openai/gpt-oss-20b`. Wenn die Basis-URL leer bleibt, nutzt die App automatisch `https://api.groq.com/openai/v1`.

### Google Gemini und Ollama
Gemini nutzt automatisch `https://generativelanguage.googleapis.com/v1beta/openai/`. Ollama nutzt typischerweise `http://localhost:11434/v1` und kann als API-Key `ollama` verwenden.
""")
            btn_reload_prompts = gr.Button("Prompts neu laden")
            prompt_status = gr.Markdown("")

        with gr.Tab("2 - Bewerter"):
            with gr.Tab("2.1 Ground Truth - Linguistic Bias"):
                gr.Markdown("**Forschungsfrage:** Wie gut stimmen mehrere LLMs mit der eigenen annotierten Ground Truth fuer Linguistic Bias ueberein? Prompt: `prompts/bewerter_linguistic_bias.txt`.")
                gt_headlines = gr.Textbox(label="Headlines", lines=8, placeholder="Eine Headline pro Zeile")
                btn_gt_table = gr.Button("Annotationstabelle erstellen")
                gt_state = gr.State([])
                gt_annotations = gr.State({})
                gt_rows, gt_mds, gt_radios = [], [], []
                for i in range(MAX_MINI):
                    with gr.Row(visible=False) as row:
                        md = gr.Markdown("")
                        radio = gr.Radio(["Low", "Medium", "High"], value="Low", label=f"Mein Label #{i + 1}")
                    gt_rows.append(row)
                    gt_mds.append(md)
                    gt_radios.append(radio)
                btn_gt_save = gr.Button("Labels speichern")
                gt_save_status = gr.Markdown("")
                state_21_results = gr.State(None)
                ms_21 = gr.CheckboxGroup(choices=llm_namen(), value=llm_namen(), label="LLMs")
                with gr.Row():
                    btn_21 = gr.Button("Bewertung starten", variant="primary")
                    btn_cancel_21 = gr.Button("Analyse abbrechen")
                status_21 = gr.Markdown("")
                table_21_comp = gr.Dataframe(label="Vergleichstabelle", interactive=False)
                table_21_kappa = gr.Dataframe(label="Cohen's Kappa", interactive=False)
                table_21_metrics = gr.Dataframe(label="Precision / Recall / F1", interactive=False)
                table_21_cm = gr.Dataframe(label="Confusion Matrices", interactive=False)
                table_21_mistakes = gr.Dataframe(label="Fehlklassifikationen", interactive=False)
                btn_export_21 = gr.Button("Export CSV + Excel")
                file_export_21 = gr.File(label="Export", file_count="multiple")

            with gr.Tab("2.2 Externe Validierung - Lyu et al."):
                gr.Markdown("**Forschungsfrage:** Wie gut erkennen LLMs Hyperpartisanship gegen den Lyu et al. (2024) Datensatz? Prompt: `prompts/bewerter_hyperpartisan.txt`.")
                lyu_file = gr.File(label="CSV hochladen", file_types=[".csv"])
                lyu_preview = gr.Dataframe(label="Vorschau", interactive=False)
                lyu_upload_status = gr.Markdown("")
                state_22_results = gr.State(None)
                with gr.Row():
                    lyu_text_col = gr.Dropdown(choices=[], label="Textspalte")
                    lyu_label_col = gr.Dropdown(choices=[], label="Labelspalte")
                lyu_n = gr.Slider(10, 200, value=50, step=10, label="Stichprobengroesse")
                ms_22 = gr.CheckboxGroup(choices=llm_namen(), value=llm_namen(), label="LLMs")
                with gr.Row():
                    btn_22 = gr.Button("Validierung starten", variant="primary")
                    btn_cancel_22 = gr.Button("Analyse abbrechen")
                status_22 = gr.Markdown("")
                table_22_comp = gr.Dataframe(label="Vergleichstabelle", interactive=False)
                table_22_metrics = gr.Dataframe(label="Precision / Recall / F1", interactive=False)
                table_22_cm = gr.Dataframe(label="Confusion Matrices", interactive=False)
                table_22_mistakes = gr.Dataframe(label="Fehlklassifikationen", interactive=False)
                btn_export_22 = gr.Button("Export CSV + Excel")
                file_export_22 = gr.File(label="Export", file_count="multiple")

        with gr.Tab("3 - Reformulierer"):
            with gr.Tab("3.1 Linguistic Bias"):
                gr.Markdown("**Forschungsfrage:** Reduzieren LLM-Reformulierungen Linguistic Bias, ohne die Bedeutung stark zu veraendern?")
                mode_31 = gr.Radio(["Ergebnisse aus Analyse 2.1 übernehmen", "Neue Headlines eingeben und direkt bewerten"], value="Ergebnisse aus Analyse 2.1 übernehmen", label="Eingabe-Modus")
                input_31 = gr.Textbox(label="Neue Headlines fuer Modus B", lines=6)
                threshold_31 = gr.Slider(0, 12, value=4, step=1, label="Schwellenwert fuer Reformulierung", info="0 = alle reformulieren. 4 = Medium + High. 8 = nur stark biased Headlines.")
                ms_31 = gr.CheckboxGroup(choices=state_models_for("bewerter_2_1") or llm_namen(), value=state_models_for("bewerter_2_1") or llm_namen(), label="LLMs")
                preview_31 = gr.Markdown(threshold_preview("bewerter_2_1", 4, state_models_for("bewerter_2_1") or llm_namen()))
                with gr.Row():
                    btn_31 = gr.Button("Reformulierung starten", variant="primary")
                    btn_cancel_31 = gr.Button("Analyse abbrechen")
                status_31 = gr.Markdown("")
                table_31_detail = gr.Dataframe(label="Vergleichstabelle", interactive=False)
                table_31_summary = gr.Dataframe(label="Zusammenfassung pro LLM", interactive=False)
                btn_export_31 = gr.Button("Export CSV + Excel")
                file_export_31 = gr.File(label="Export", file_count="multiple")

            with gr.Tab("3.2 Hyperpartisanship"):
                gr.Markdown("**Forschungsfrage:** Reduzieren LLM-Reformulierungen Hyperpartisanship im Lyu-Kontext?")
                mode_32 = gr.Radio(["Ergebnisse aus Analyse 2.2 übernehmen", "Neue CSV laden und direkt bewerten"], value="Ergebnisse aus Analyse 2.2 übernehmen", label="Eingabe-Modus")
                file_32 = gr.File(label="CSV fuer Modus B", file_types=[".csv"])
                preview_32_csv = gr.Dataframe(label="CSV-Vorschau", interactive=False)
                upload_32_status = gr.Markdown("")
                with gr.Row():
                    text_32_col = gr.Dropdown(choices=[], label="Textspalte")
                    label_32_col = gr.Dropdown(choices=[], label="Labelspalte")
                n_32 = gr.Slider(10, 200, value=50, step=10, label="Stichprobengroesse fuer Modus B")
                threshold_32 = gr.Slider(0, 12, value=4, step=1, label="Schwellenwert fuer Reformulierung", info="0 = alle reformulieren. 4 = Medium + High. 8 = nur stark biased Headlines.")
                only_positive_32 = gr.Checkbox(label="Nur Headlines mit Lyu-Label = hyperpartisan (1) reformulieren", value=False)
                ms_32 = gr.CheckboxGroup(choices=state_models_for("bewerter_2_2") or llm_namen(), value=state_models_for("bewerter_2_2") or llm_namen(), label="LLMs")
                preview_32 = gr.Markdown(threshold_preview("bewerter_2_2", 4, state_models_for("bewerter_2_2") or llm_namen()))
                with gr.Row():
                    btn_32 = gr.Button("Reformulierung starten", variant="primary")
                    btn_cancel_32 = gr.Button("Analyse abbrechen")
                status_32 = gr.Markdown("")
                table_32_detail = gr.Dataframe(label="Vergleichstabelle", interactive=False)
                table_32_summary = gr.Dataframe(label="Zusammenfassung pro LLM", interactive=False)
                btn_export_32 = gr.Button("Export CSV + Excel")
                file_export_32 = gr.File(label="Export", file_count="multiple")

        with gr.Tab("4 - Dashboard & Export"):
            gr.Markdown("Aggregiert die zuletzt berechneten Ergebnisse aus Tab 2 und Tab 3.")
            btn_dashboard = gr.Button("Aktualisieren")
            dash_21 = gr.Dataframe(label="Analyse 2.1 Metriken", interactive=False)
            dash_22 = gr.Dataframe(label="Analyse 2.2 Metriken", interactive=False)
            dash_31 = gr.Dataframe(label="Analyse 3.1 Zusammenfassung", interactive=False)
            dash_32 = gr.Dataframe(label="Analyse 3.2 Zusammenfassung", interactive=False)
            btn_export_all = gr.Button("Alle Ergebnisse exportieren", variant="primary")
            files_all = gr.File(label="Export-Dateien", file_count="multiple")
            status_all = gr.Markdown("")

    btn_save.click(llm_hinzufuegen, inputs=[tf_name, dd_provider, tf_model_id, tf_api_key, tf_base_url], outputs=[llm_table, cfg_status, ms_21, ms_22, ms_31, ms_32, dd_remove])
    btn_remove.click(llm_entfernen, inputs=[dd_remove], outputs=[llm_table, cfg_status, ms_21, ms_22, ms_31, ms_32, dd_remove])
    btn_reload_prompts.click(reload_prompts, outputs=[prompt_status])

    btn_gt_table.click(mini_tabelle_erstellen, inputs=[gt_headlines], outputs=[gt_state] + gt_rows + gt_mds + gt_radios)
    btn_gt_save.click(labels_speichern, inputs=[gt_state] + gt_radios, outputs=[gt_annotations, gt_save_status])
    run_21 = btn_21.click(analyse_2_1_logged, inputs=[gt_annotations, ms_21], outputs=[table_21_comp, table_21_kappa, table_21_metrics, table_21_cm, table_21_mistakes, state_21_results, ms_31, preview_31, status_21])
    btn_cancel_21.click(lambda: "Abbruch angefordert. Der aktuelle API-Call kann noch kurz fertig laufen.", outputs=[status_21], cancels=[run_21])
    btn_export_21.click(export_2_1, outputs=[file_export_21])

    lyu_file.change(lyu_csv_vorschau, inputs=[lyu_file], outputs=[lyu_preview, lyu_text_col, lyu_label_col, lyu_upload_status])
    run_22 = btn_22.click(analyse_2_2_logged, inputs=[lyu_file, lyu_text_col, lyu_label_col, lyu_n, ms_22], outputs=[table_22_comp, table_22_metrics, table_22_cm, table_22_mistakes, state_22_results, ms_32, preview_32, status_22])
    btn_cancel_22.click(lambda: "Abbruch angefordert. Der aktuelle API-Call kann noch kurz fertig laufen.", outputs=[status_22], cancels=[run_22])
    btn_export_22.click(export_2_2, outputs=[file_export_22])

    mode_31.change(lambda m: update_mode_models(m, "bewerter_2_1"), inputs=[mode_31], outputs=[ms_31])
    threshold_31.change(lambda t, m: threshold_preview("bewerter_2_1", t, m), inputs=[threshold_31, ms_31], outputs=[preview_31])
    ms_31.change(lambda m, t: threshold_preview("bewerter_2_1", t, m), inputs=[ms_31, threshold_31], outputs=[preview_31])
    run_31 = btn_31.click(analyse_3_1_logged, inputs=[mode_31, input_31, ms_31, threshold_31, state_21_results], outputs=[table_31_detail, table_31_summary, status_31])
    btn_cancel_31.click(lambda: "Abbruch angefordert. Der aktuelle API-Call kann noch kurz fertig laufen.", outputs=[status_31], cancels=[run_31])
    btn_export_31.click(export_3_1, outputs=[file_export_31])

    file_32.change(lyu_csv_vorschau, inputs=[file_32], outputs=[preview_32_csv, text_32_col, label_32_col, upload_32_status])
    mode_32.change(lambda m: update_mode_models(m, "bewerter_2_2"), inputs=[mode_32], outputs=[ms_32])
    threshold_32.change(lambda t, m, op: threshold_preview("bewerter_2_2", t, m, op), inputs=[threshold_32, ms_32, only_positive_32], outputs=[preview_32])
    ms_32.change(lambda m, t, op: threshold_preview("bewerter_2_2", t, m, op), inputs=[ms_32, threshold_32, only_positive_32], outputs=[preview_32])
    only_positive_32.change(lambda op, t, m: threshold_preview("bewerter_2_2", t, m, op), inputs=[only_positive_32, threshold_32, ms_32], outputs=[preview_32])
    run_32 = btn_32.click(analyse_3_2_logged, inputs=[mode_32, file_32, text_32_col, label_32_col, n_32, ms_32, threshold_32, only_positive_32, state_22_results], outputs=[table_32_detail, table_32_summary, status_32])
    btn_cancel_32.click(lambda: "Abbruch angefordert. Der aktuelle API-Call kann noch kurz fertig laufen.", outputs=[status_32], cancels=[run_32])
    btn_export_32.click(export_3_2, outputs=[file_export_32])

    btn_dashboard.click(dashboard_refresh, outputs=[dash_21, dash_22, dash_31, dash_32])
    btn_export_all.click(export_all, outputs=[files_all, status_all])


if __name__ == "__main__":
    app.launch()
